import openpyxl
# (*** 수정: 열 문자 <-> 숫자 변환을 위해 임포트 ***)
from openpyxl.utils import get_column_letter, column_index_from_string
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
import io
import os
import re
from collections import Counter
from flask_sqlalchemy import SQLAlchemy
# (*** 수정: 정렬을 위해 desc 임포트 ***)
from sqlalchemy import or_, func, text, Integer, String, update, exc, delete, desc
from sqlalchemy.orm import joinedload, aliased
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'sqlite:///' + os.path.join(app.root_path, 'database.db') # 기본 SQLite 경로
)
app.config['SECRET_KEY'] = 'wasabi-check-secret-key'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = '/tmp'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

db = SQLAlchemy(app)

# (*** 중요: 배포 시 타임아웃을 막기 위해 @app.before_request 유지 ***)
# 앱이 시작되고 "첫 번째 요청"이 들어왔을 때 DB를 생성합니다.
@app.before_request
def create_tables_if_not_exist():
    if not getattr(app, '_tables_checked', False):
        print("Checking if tables exist...")
        with app.app_context():
            db.create_all()
        app._tables_checked = True
        print("Tables check complete.")


IMAGE_URL_PREFIX = 'https://files.ebizway.co.kr/files/10249/Style/'

@app.context_processor
def inject_image_url_prefix():
    return dict(IMAGE_URL_PREFIX=IMAGE_URL_PREFIX)

def clean_string_upper(s, default=''):
    if not (s is not None and s == s): return default
    return str(s).replace('-', '').replace(' ', '').strip().upper()

CHOSUNG_LIST = ['ㄱ', 'ㄲ', 'ㄴ', 'ㄷ', 'ㄸ', 'ㄹ', 'ㅁ', 'ㅂ', 'ㅃ', 'ㅅ', 'ㅆ', 'ㅇ', 'ㅈ', 'ㅉ', 'ㅊ', 'ㅋ', 'ㅌ', 'ㅍ', 'ㅎ']
def get_choseong(text):
    if not (text is not None and text == text): return ''
    text_cleaned = str(text).replace('-', '').replace(' ', '').strip().upper()
    result = ''
    for char in text_cleaned:
        if '가' <= char <= '힣':
            code = ord(char) - 0xAC00
            cho_index = code // (21 * 28)
            result += CHOSUNG_LIST[cho_index]
        elif 'ㄱ' <= char <= 'ㅎ':
            result += char
        elif 'A' <= char <= 'Z' or '0' <= char <= '9':
            result += char
    return result

class Product(db.Model):
    __tablename__ = 'products'
    product_number = db.Column(String, primary_key=True)
    product_name = db.Column(String, nullable=False)
    is_favorite = db.Column(Integer, default=0)
    
    # (*** [최적화 2] 적용됨: DB 인덱스 ***)
    release_year = db.Column(Integer, nullable=True, index=True)
    item_category = db.Column(String, nullable=True, index=True)
    
    product_number_cleaned = db.Column(String, index=True)
    product_name_cleaned = db.Column(String, index=True)
    product_name_choseong = db.Column(String, index=True) 
    variants = db.relationship('Variant', backref='product', lazy=True, cascade="all, delete-orphan")

class Variant(db.Model):
    __tablename__ = 'variants'
    barcode = db.Column(String, primary_key=True)
    product_number = db.Column(String, db.ForeignKey('products.product_number'), nullable=False)
    color = db.Column(String)
    size = db.Column(String)
    store_stock = db.Column(Integer, default=0)
    hq_stock = db.Column(Integer, default=0)
    original_price = db.Column(Integer, default=0)
    sale_price = db.Column(Integer, default=0)
    actual_stock = db.Column(Integer, nullable=True)
    barcode_cleaned = db.Column(String, index=True, unique=True)
    color_cleaned = db.Column(String, index=True)
    size_cleaned = db.Column(String, index=True)

def init_db():
    with app.app_context():
        db.create_all()
        print("DB 테이블 초기화/검증 완료.")


def generate_barcode(row_data):
    try:
        pn = str(row_data.get('product_number', '')).strip()
        color = str(row_data.get('color', '')).strip()
        size = str(row_data.get('size', '')).strip()
        pn_cleaned = pn.replace('-', ''); size_upper = size.upper()
        pn_final = pn_cleaned + '00' if len(pn_cleaned) <= 10 else pn_cleaned
        if size_upper == 'FREE': size_final = '00F'
        elif size.isdigit() and len(size) == 2: size_final = '0' + size
        elif size.isalpha() and len(size) == 2: size_final = '0' + size_upper
        else: size_final = size
        if pn_final and color and size_final: return f"{pn_final}{color}{size_final}".upper()
        else: print(f"Barcode generation skipped: {row_data}"); return None
    except Exception as e: print(f"Error generating barcode for {row_data}: {e}"); return None

# (*** 수정: 열 문자를 기반으로 엑셀을 읽는 새 헬퍼 함수 ***)
def _read_excel_by_letter_to_dicts(file_bytes, col_map, dtype_map):
    """
    엑셀 파일을 열고, 지정된 열 문자(A, B, C...)를 기준으로 데이터를 읽어
    [{'key1': 'valA', 'key2': 'valB'}, ...] 형태의 딕셔너리 리스트로 반환합니다.
    
    :param file_bytes: 엑셀 파일의 바이트 데이터
    :param col_map: {'product_number': 'A', 'product_name': 'C', ...} 형태의 맵
    :param dtype_map: {'product_number': str, ...} 형태의 타입 맵
    :return: 데이터 딕셔너리 리스트
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    # {'product_number': 'A'} -> {'product_number': 1}
    col_idx_map = {}
    for key, col_letter in col_map.items():
        if col_letter:
            try:
                col_idx_map[key] = column_index_from_string(col_letter.strip().upper())
            except Exception:
                raise ValueError(f"'{col_letter}'는 유효한 엑셀 열 문자가 아닙니다.")
        else:
            raise ValueError(f"'{key}'에 대한 열 문자가 선택되지 않았습니다.")

    data_rows = []
    # 2번째 행부터 데이터로 읽음 (1행은 헤더로 간주)
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        all_none = True
        
        for key, col_idx in col_idx_map.items():
            # openpyxl의 row 객체는 0-based 인덱싱을 사용하므로 1을 뺌
            cell_val = row[col_idx - 1].value
            
            if cell_val is None:
                cell_val = ''
            
            # 지정된 dtype이 str이면 강제로 문자열 변환
            if key in dtype_map and dtype_map[key] == str:
                cell_val = str(cell_val)
                
            if cell_val != '':
                all_none = False
            
            row_data[key] = cell_val
            
        if not all_none:
            data_rows.append(row_data)
            
    return data_rows

# (*** 삭제: 기존 _read_excel_to_dicts 함수 (더 이상 사용 안 함) ***)
# def _read_excel_to_dicts(file_bytes, dtype_map): ...

@app.route('/import_excel', methods=['POST'])
def import_excel():
    redirect_url = url_for('stock_management')
    if 'excel_file' not in request.files: flash('파일 선택 안됨.', 'error'); return redirect(redirect_url)
    file = request.files['excel_file']
    if file.filename == '': flash('파일 선택 안됨.', 'error'); return redirect(redirect_url)
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            file_bytes = file.read()
            
            # (*** 중요: DB 업로드는 기존 헤더 이름 방식 유지를 위해 _read_excel_to_dicts 재구현 ***)
            # (이 함수는 /import_excel 에서만 사용되도록 간단히 재정의)
            def _read_excel_by_header(file_bytes_inner, dtype_map):
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes_inner), data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                header_map = {h: i for i, h in enumerate(headers) if h}
                data_rows = []
                for row in ws.iter_rows(min_row=2):
                    row_data = {}
                    all_none = True
                    for header, col_idx in header_map.items():
                        cell_val = row[col_idx].value
                        if cell_val is None: cell_val = ''
                        if header in dtype_map and dtype_map[header] == str:
                            cell_val = str(cell_val)
                        if cell_val != '': all_none = False
                        row_data[header] = cell_val
                    if not all_none: data_rows.append(row_data)
                return data_rows, header_map.keys()

            dtype_map = {'product_number': str, 'color': str, 'size': str, 'product_name': str}
            data, df_columns = _read_excel_by_header(file_bytes, dtype_map)
            
            required_cols = ['product_number', 'product_name', 'color', 'size', 'release_year', 'item_category', 'original_price', 'sale_price', 'store_stock', 'hq_stock', 'actual_stock', 'is_favorite']
            if not all(col in df_columns for col in required_cols):
                 missing = [col for col in required_cols if col not in df_columns]; flash(f"엑셀 컬럼명 오류. 누락: {missing}", 'error'); return redirect(redirect_url)

            generated_barcodes = []
            valid_rows = []
            for row in data:
                row['generated_barcode'] = generate_barcode(row)
                if row['generated_barcode']:
                    generated_barcodes.append(row['generated_barcode'])
                    valid_rows.append(row)
                else:
                    print(f"Barcode generation failed for row: {row.get('product_number')}")
            
            if len(valid_rows) != len(data):
                flash(f"바코드 생성 실패: {len(data) - len(valid_rows)}개 행 누락.", 'error'); 
            
            duplicates = [item for item, count in Counter(generated_barcodes).items() if count > 1]
            if duplicates:
                flash(f"바코드 생성 실패: 중복 발생 ({', '.join(duplicates[:3])}...).", 'error'); return redirect(redirect_url)

            if not valid_rows: flash("처리할 유효 데이터 없음.", 'error'); return redirect(redirect_url)

            db.session.query(Variant).delete(); db.session.query(Product).delete(); db.session.commit()

            products_data = {}
            variants_data = []
            
            for row in valid_rows:
                pn = row['product_number']
                if not pn: continue

                if pn not in products_data:
                    try:
                        year = int(row.get('release_year')) if str(row.get('release_year')).isdigit() else None
                    except (ValueError, TypeError):
                        year = None
                    
                    product_name = str(row.get('product_name', '')) or f"{pn} (신규)"
                    
                    products_data[pn] = {
                        'product_number': pn,
                        'product_name': product_name,
                        'release_year': year,
                        'item_category': str(row.get('item_category', '')) or None,
                        'is_favorite': int(row.get('is_favorite', 0) or 0),
                        'product_number_cleaned': clean_string_upper(pn),
                        'product_name_cleaned': clean_string_upper(product_name),
                        'product_name_choseong': get_choseong(product_name)
                    }

                try:
                    actual_stock = int(row.get('actual_stock')) if str(row.get('actual_stock')).isdigit() else None
                except (ValueError, TypeError):
                    actual_stock = None
                
                variants_data.append({
                    'barcode': row['generated_barcode'],
                    'product_number': pn,
                    'color': str(row.get('color', '')),
                    'size': str(row.get('size', '')),
                    'store_stock': int(row.get('store_stock', 0) or 0),
                    'hq_stock': int(row.get('hq_stock', 0) or 0),
                    'original_price': int(row.get('original_price', 0) or 0),
                    'sale_price': int(row.get('sale_price', 0) or 0),
                    'actual_stock': actual_stock,
                    'barcode_cleaned': clean_string_upper(row['generated_barcode']),
                    'color_cleaned': clean_string_upper(row.get('color', '')),
                    'size_cleaned': clean_string_upper(row.get('size', ''))
                })

            db.session.bulk_insert_mappings(Product, products_data.values())
            db.session.bulk_insert_mappings(Variant, variants_data)

            db.session.commit()
            flash(f"성공 ({file.filename}): {len(products_data)}개 상품, {len(variants_data)}개 SKU 임포트 (바코드 자동 생성).", 'success')
        except Exception as e:
            db.session.rollback(); flash(f"임포트 오류: {e}", 'error'); print(f"Import error: {e}"); import traceback; traceback.print_exc()
        return redirect(redirect_url)
    else: flash('엑셀 파일만 가능.', 'error'); return redirect(redirect_url)

def _write_dicts_to_excel(data_list, column_order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DB_Export'
    
    ws.append(column_order)
    
    for row_data in data_list:
        row_to_append = [row_data.get(col_name, '') for col_name in column_order]
        ws.append(row_to_append)
        
    for i, col_letter in enumerate(get_column_letter(idx) for idx in range(1, len(column_order) + 1)):
        ws.column_dimensions[col_letter].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/export_db_excel')
def export_db_excel():
    try:
        results = db.session.query(
            Product.product_number,
            Product.product_name,
            Variant.color,
            Variant.size,
            Product.release_year,
            Product.item_category,
            Variant.original_price,
            Variant.sale_price,
            Variant.store_stock,
            Variant.hq_stock,
            Variant.actual_stock,
            Product.is_favorite,
            Variant.barcode
        ).join(Variant, Product.product_number == Variant.product_number).order_by(
            Product.product_number, Variant.color, Variant.size
        ).all()

        if not results:
            flash("DB에 데이터가 없습니다.", "warning")
            return redirect(url_for('stock_management'))

        column_order = [
            'product_number', 'product_name', 'color', 'size',
            'release_year', 'item_category', 'original_price', 'sale_price',
            'store_stock', 'hq_stock', 'actual_stock', 'is_favorite', 'barcode'
        ]
        
        data_list = []
        for row in results:
            data_list.append({
                'product_number': row.product_number,
                'product_name': row.product_name,
                'color': row.color,
                'size': row.size,
                'release_year': '' if row.release_year is None else row.release_year,
                'item_category': '' if row.item_category is None else row.item_category,
                'original_price': row.original_price,
                'sale_price': row.sale_price,
                'store_stock': row.store_stock,
                'hq_stock': row.hq_stock,
                'actual_stock': '' if row.actual_stock is None else row.actual_stock,
                'is_favorite': row.is_favorite,
                'barcode': row.barcode
            })

        output = _write_dicts_to_excel(data_list, column_order)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='honeyflow_db_export.xlsx'
        )
    except Exception as e:
        flash(f"DB 엑셀 출력 오류: {e}", 'error')
        print(f"DB Export error: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('stock_management'))


@app.route('/')
def index():
    query_param = request.args.get('query', '')
    category_param = request.args.get('selected_category', '전체')
    showing_favorites = False

    base_query = Product.query.options(joinedload(Product.variants))

    if query_param:
        search_term_cleaned = clean_string_upper(query_param)
        search_like = f'%{search_term_cleaned}%'
        base_query = base_query.filter(
            or_(
                Product.product_number_cleaned.like(search_like),
                Product.product_name_cleaned.like(search_like),
                Product.product_name_choseong.like(search_like)
            )
        )

    if category_param and category_param != '전체':
        base_query = base_query.filter(Product.item_category == category_param)

    if query_param or (category_param and category_param != '전체'):
        # (*** 수정: 검색 결과 정렬 순서를 년도(최신) -> 이름순으로 변경 ***)
        products = base_query.order_by(Product.release_year.desc(), Product.product_name).all()
    else:
        showing_favorites = True
        products = base_query.filter(Product.is_favorite == 1).order_by(Product.item_category, Product.product_name).all()

    return render_template('index.html',
                           products=products,
                           query=query_param,
                           showing_favorites=showing_favorites,
                           active_page='index',
                           selected_category=category_param)

@app.route('/api/live_search', methods=['POST'])
def live_search():
    data = request.json
    query_param = data.get('query', '')
    category_param = data.get('category', '전체')

    base_query = Product.query.options(joinedload(Product.variants))
    showing_favorites = False

    is_searching = bool(query_param) or (category_param and category_param != '전체')

    if is_searching:
        if query_param:
            search_term_cleaned = clean_string_upper(query_param)
            search_like = f'%{search_term_cleaned}%'
            base_query = base_query.filter(
                or_(
                    Product.product_number_cleaned.like(search_like),
                    Product.product_name_cleaned.like(search_like),
                    Product.product_name_choseong.like(search_like)
                )
            )

        if category_param and category_param != '전체':
            base_query = base_query.filter(Product.item_category == category_param)

        # (*** 수정: 실시간 검색 결과 정렬 순서를 년도(최신) -> 이름순으로 변경 ***)
        products = base_query.order_by(Product.release_year.desc(), Product.product_name).all()
    else:
        showing_favorites = True
        products = base_query.filter(Product.is_favorite == 1).order_by(Product.item_category, Product.product_name).all()

    results_list = []
    for product in products:
        image_pn = product.product_number.split(' ')[0]

        colors = ""
        sale_price_f = "가격정보없음"
        original_price_f = 0
        discount_f = "-"

        if product.variants:
            colors_list = sorted(list(set(v.color for v in product.variants if v.color)))
            colors = ", ".join(colors_list)
            first_variant = product.variants[0]
            sale_price_f = f"{first_variant.sale_price:,d}원"
            original_price_f = first_variant.original_price
            if original_price_f and original_price_f > 0:
                discount_f = f"{int((1 - (first_variant.sale_price / original_price_f)) * 100)}%"
            else:
                discount_f = "0%"

        results_list.append({
            "product_number": product.product_number,
            "product_name": product.product_name,
            "image_pn": image_pn,
            "colors": colors,
            "sale_price": sale_price_f,
            "original_price": original_price_f,
            "discount": discount_f
        })

    return jsonify({
        "status": "success",
        "products": results_list,
        "showing_favorites": showing_favorites,
        "selected_category": category_param
    })


# (*** [사용자 요청] 추가: 필터 옵션을 조회하는 헬퍼 함수 ***)
def get_filter_options():
    """DB에서 필터링 옵션을 조회"""
    try:
        # 1. 품목 (인덱스 사용으로 빠름)
        categories = [r[0] for r in db.session.query(Product.item_category).distinct().order_by(Product.item_category).all() if r[0]]
        
        # 2. 출시년도 (인덱스 사용으로 빠름)
        years = [r[0] for r in db.session.query(Product.release_year).distinct().order_by(Product.release_year.desc()).all() if r[0]]
        
        # 3. 색상
        colors = [r[0] for r in db.session.query(Variant.color).distinct().order_by(Variant.color).all() if r[0]]
        
        # 4. 사이즈 (복잡한 정렬 필요)
        sizes_raw = [r[0] for r in db.session.query(Variant.size).distinct().all() if r[0]]
        
        def size_sort_key(size_str):
            """사이즈 문자열(예: 100, L, M, 2XL)을 정렬 가능한 키로 변환"""
            size_str_upper = str(size_str).upper().strip()
            custom_order = {'2XS': 'XXS', '2XL': 'XXL', '3XL': 'XXXL'}
            size_str_upper = custom_order.get(size_str_upper, size_str_upper)
            order_map = {'XXS': 0, 'XS': 1, 'S': 2, 'M': 3, 'L': 4, 'XL': 5, 'XXL': 6, 'XXXL': 7}
            
            if size_str_upper.isdigit():
                return (1, int(size_str_upper)) # 숫자
            elif size_str_upper in order_map:
                return (2, order_map[size_str_upper]) # 표준 의류 사이즈
            else:
                return (3, size_str_upper) # 기타 문자
        
        sizes = sorted(sizes_raw, key=size_sort_key)

        # (*** 수정: 최초가 및 판매가 목록 추가 (0원 이상) ***)
        original_prices = [r[0] for r in db.session.query(Variant.original_price).distinct().order_by(Variant.original_price.desc()).all() if r[0] and r[0] > 0]
        sale_prices = [r[0] for r in db.session.query(Variant.sale_price).distinct().order_by(Variant.sale_price.desc()).all() if r[0] and r[0] > 0]


        return {
            'categories': categories,
            'years': years,
            'colors': colors,
            'sizes': sizes,
            'original_prices': original_prices, # (*** 수정: 추가 ***)
            'sale_prices': sale_prices          # (*** 수정: 추가 ***)
        }
    except Exception as e:
        print(f"Error fetching filter options: {e}")
        return { 'categories': [], 'years': [], 'colors': [], 'sizes': [], 'original_prices': [], 'sale_prices': [] }


@app.route('/list')
def list_page():
    try:
        # (*** [사용자 요청] 추가: 필터 옵션 조회 ***)
        filter_options = get_filter_options()

        # (*** [최적화 1] 적용됨: .paginate() 사용 ***)
        page = request.args.get('page', 1, type=int)
        per_page = 20  # 한 페이지에 20개씩 표시
        
        pagination = Product.query.options(joinedload(Product.variants)).order_by(Product.item_category, Product.product_name).paginate(page=page, per_page=per_page, error_out=False)
        
        products = pagination.items
        
        return render_template('list.html',
                               products=products, 
                               pagination=pagination, # pagination 객체 전달
                               query="전체 목록", 
                               showing_all=True, 
                               active_page='list', 
                               advanced_search_params={}, # 상세 검색 파라미터가 없으므로 빈 dict 전달
                               filter_options=filter_options) # (*** [사용자 요청] 추가 ***)
    except Exception as e: 
        flash(f"목록 조회 오류: {e}", 'error')
        return redirect(url_for('index'))

@app.route('/advanced_search')
def advanced_search():
    try:
        # (*** [사용자 요청] 추가: 필터 옵션 조회 ***)
        filter_options = get_filter_options()

        # (*** [최적화 1] 적용됨: 페이지네이션 파라미터 ***)
        page = request.args.get('page', 1, type=int)
        per_page = 20 

        query = Product.query.options(joinedload(Product.variants)).join(Product.variants)
        
        # (*** [최적화 1] 적용됨: 'page'를 제외한 검색 파라미터 ***)
        params = request.args.copy()
        params.pop('page', None) 
        
        search_active = False
        query_summary_parts = [] # 검색 조건 요약을 위해 유지

        # --- 텍스트 필드 검색 (기존 로직) ---
        if params.get('product_number'):
            value = clean_string_upper(params.get('product_number'));
            query = query.filter(Product.product_number_cleaned.like(f"%{value}%"));
            search_active = True; query_summary_parts.append(f"품번: {params.get('product_number')}")
        if params.get('product_name'):
            value_like = f"%{clean_string_upper(params.get('product_name'))}%"
            query = query.filter(
                or_(
                    Product.product_name_cleaned.like(value_like),
                    Product.product_name_choseong.like(value_like)
                )
            )
            search_active = True; query_summary_parts.append(f"품명: {params.get('product_name')}")
        
        # --- 드롭다운 필드 검색 (정확히 일치) ---
        if params.get('color'):
            value = params.get('color')
            query = query.filter(Variant.color == value) # (LIKE -> ==)
            search_active = True; query_summary_parts.append(f"색상: {value}")
        if params.get('size'):
            value = params.get('size')
            query = query.filter(Variant.size == value) # (LIKE -> ==)
            search_active = True; query_summary_parts.append(f"사이즈: {value}")
        if params.get('item_category'):
            value = params.get('item_category')
            query = query.filter(Product.item_category == value) # (LIKE -> ==)
            search_active = True; query_summary_parts.append(f"품목: {value}")
        if params.get('release_year'):
            try: 
                year = int(params.get('release_year'))
                query = query.filter(Product.release_year == year) # (LIKE -> ==)
                search_active = True; query_summary_parts.append(f"년도: {year}")
            except ValueError: 
                pass # 값이 있지만 숫자가 아니면 무시

        # --- (*** 수정: 가격/할인율 필드 검색 ***) ---
        
        # (*** 삭제: 기존 Min/Max 가격 로직 ***)
        
        # (*** 추가: 새로운 드롭다운 가격 로직 (정확히 일치) ***)
        if params.get('original_price'):
            try: 
                price = int(params.get('original_price'))
                query = query.filter(Variant.original_price == price)
                search_active = True; query_summary_parts.append(f"최초가: {price:,.0f}원")
            except ValueError: 
                pass 
        if params.get('sale_price'):
            try: 
                price = int(params.get('sale_price'))
                query = query.filter(Variant.sale_price == price)
                search_active = True; query_summary_parts.append(f"판매가: {price:,.0f}원")
            except ValueError: 
                pass

        # (*** 유지: 할인율 로직 ***)
        if params.get('min_discount'):
            try:
                discount = int(params.get('min_discount'))
                if discount > 0: ratio = 1.0 - (discount / 100.0); query = query.filter(Variant.original_price > 0).filter(Variant.sale_price <= (Variant.original_price * ratio)); search_active = True; query_summary_parts.append(f"할인율: {discount}% 이상")
            except ValueError: flash(f"할인율 오류: '{params.get('min_discount')}' 유효숫자 아님.", 'warning')


        if not search_active: 
            products = []
            pagination = None 
            query_summary = "상세 검색: 조건 없음"
        else: 
            # (*** [최적화 1] 적용됨: .paginate() 사용 ***)
            pagination = query.distinct().order_by(Product.product_name).paginate(page=page, per_page=per_page, error_out=False)
            products = pagination.items
            query_summary = f"상세 검색: {', '.join(query_summary_parts)}"

        return render_template( 'list.html', 
                               products=products, 
                               pagination=pagination, 
                               query=query_summary, 
                               showing_all=False, 
                               active_page='list', 
                               advanced_search_params=params,
                               filter_options=filter_options) # (*** [사용자 요청] 추가 ***)
    except Exception as e: 
        flash(f"검색 오류: {e}", 'error')
        return redirect(url_for('index'))

@app.route('/stock')
def stock_management():
    missing_data_products = Product.query.filter(or_(Product.release_year.is_(None), Product.item_category.is_(None))).order_by(Product.product_number).all()
    return render_template('stock.html', active_page='stock', missing_data_products=missing_data_products)

@app.route('/reset_actual_stock', methods=['POST'])
def reset_actual_stock():
    try: stmt = update(Variant).values(actual_stock=None); db.session.execute(stmt); db.session.commit(); flash('실사재고 초기화 완료.', 'success')
    except Exception as e: db.session.rollback(); flash(f'초기화 오류: {e}', 'error')
    return redirect(url_for('stock_management'))

@app.route('/reset_database_completely', methods=['POST'])
def reset_database_completely():
    try:
        db.drop_all()
        db.create_all()
        flash('데이터베이스 전체 초기화 완료. (모든 데이터 삭제됨)', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'DB 전체 초기화 오류: {e}', 'error')
        print(f"DB Reset Error: {e}")
    return redirect(url_for('stock_management'))

# (*** 신규: 엑셀 분석 API ***)
@app.route('/api/analyze_excel', methods=['POST'])
def analyze_excel():
    """
    업로드된 엑셀 파일을 분석하여
    (1) 최대 열 문자 리스트 (['A', 'B', ..., 'G'])
    (2) 열별 1~5행 미리보기 데이터
    를 JSON으로 반환합니다.
    """
    if 'excel_file' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 없습니다.'}), 400
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': '파일이 선택되지 않았습니다.'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'status': 'error', 'message': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'}), 400

    try:
        file_bytes = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        # (1) 최대 열 문자 리스트 생성
        max_col_index = ws.max_column
        # (A-Z 까지만 우선 지원, 26개)
        if max_col_index > 26: max_col_index = 26 
        column_letters = [get_column_letter(i) for i in range(1, max_col_index + 1)]
        
        # (2) 열별 미리보기 데이터 (최대 5행)
        preview_data = {}
        # 1(헤더행) + 4(데이터행) = 5행
        max_row_preview = min(6, ws.max_row + 1) 
        
        if max_row_preview <= 1: # 데이터가 없는 파일
             return jsonify({'status': 'error', 'message': '파일에 데이터가 없습니다.'}), 400

        for col_letter in column_letters:
            col_data = []
            col_index = column_index_from_string(col_letter)
            for i in range(1, max_row_preview):
                cell_val = ws.cell(row=i, column=col_index).value
                col_data.append(str(cell_val) if cell_val is not None else '')
            preview_data[col_letter] = col_data
            
        return jsonify({
            'status': 'success',
            'column_letters': column_letters,
            'preview_data': preview_data
        })
        
    except Exception as e:
        print(f"Excel analyze error: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'엑셀 파일 분석 중 오류 발생: {e}'}), 500


def _process_stock_update_excel(file, form_data, stock_type):
    if not file or file.filename == '':
        # (stock.html의 JS가 파일 누락을 먼저 잡겠지만, 이중 방어)
        flash('파일 선택 안됨.', 'error')
        return None, None

    # (*** 수정: 열 문자를 받도록 변경 ***)
    col_pn_letter = form_data.get('col_product_number')
    col_pname_letter = form_data.get('col_product_name')
    col_color_letter = form_data.get('col_color')
    col_size_letter = form_data.get('col_size')
    col_stock_letter = form_data.get('col_stock')

    # (*** 수정: 열 문자 유효성 검사 ***)
    if not all([col_pn_letter, col_pname_letter, col_color_letter, col_size_letter, col_stock_letter]):
        flash('모든 항목(품번, 품명, 컬러, 사이즈, 재고)의 엑셀 열 문자를 선택해야 합니다.', 'error')
        return None, None

    try:
        file_bytes = file.read()
        
        # (*** 수정: 헬퍼 함수에 전달할 맵 생성 ***)
        col_map = {
            'product_number': col_pn_letter,
            'product_name': col_pname_letter,
            'color': col_color_letter,
            'size': col_size_letter,
            'stock': col_stock_letter
        }
        dtype_map = {
            'product_number': str, 
            'product_name': str, 
            'color': str, 
            'size': str
        }
        
        # (*** 수정: 열 문자를 기반으로 읽는 새 헬퍼 함수 호출 ***)
        data = _read_excel_by_letter_to_dicts(file_bytes, col_map, dtype_map)
        
        # (*** 삭제: 기존 헤더명 기반 컬럼 체크 로직 ***)
        # required_cols = [...]
        # if not all(col in df_columns for col in required_cols): ...

        
        print("Fetching all variants for lookup...")
        existing_variants = db.session.query(Variant).all()
        variant_lookup = {
            (v.product_number, v.color, v.size): v
            for v in existing_variants
        }
        print(f"Variant lookup created with {len(variant_lookup)} variants.")

        print("Fetching all products for lookup...")
        existing_products = db.session.query(Product).all()
        product_lookup = {p.product_number: p for p in existing_products}
        print(f"Product lookup created with {len(product_lookup)} products.")

        print("Creating product default data lookup...")
        product_default_data = {}
        for v in existing_variants:
            if v.product_number not in product_default_data and (v.original_price > 0 or v.sale_price > 0):
                product_default_data[v.product_number] = {
                    'original_price': v.original_price,
                    'sale_price': v.sale_price
                }
        print("Default data lookup created.")


        updated_count = 0
        added_variants = 0
        added_products = 0
        variants_to_add = []
        products_to_add = {} 

        for row in data:
            # (*** 수정: col_map에서 정의한 generic key로 데이터 추출 ***)
            pn = str(row.get('product_number', '')).strip()
            product_name_from_excel = str(row.get('product_name', '')).strip()
            color = str(row.get('color', '')).strip()
            size = str(row.get('size', '')).strip()
            stock_value_raw = row.get('stock', None)

            try:
                stock_value = int(stock_value_raw) if stock_value_raw is not None and str(stock_value_raw).isdigit() else None
            except (ValueError, TypeError):
                stock_value = None

            if not pn or stock_value is None or stock_value < 0:
                print(f"Skipping excel row: Invalid data (PN: {pn}, Name: {product_name_from_excel}, Color: {color}, Size: {size}, Stock: {stock_value_raw})")
                continue

            variant = variant_lookup.get((pn, color, size))

            if variant:
                if stock_type == 'store':
                    variant.store_stock = stock_value
                elif stock_type == 'hq':
                    variant.hq_stock = stock_value
                updated_count += 1
                
                existing_product = product_lookup.get(pn)
                if existing_product and product_name_from_excel and existing_product.product_name != product_name_from_excel:
                    print(f"Updating product name for {pn}: '{existing_product.product_name}' -> '{product_name_from_excel}'")
                    existing_product.product_name = product_name_from_excel
                    existing_product.product_name_cleaned = clean_string_upper(product_name_from_excel)
                    existing_product.product_name_choseong = get_choseong(product_name_from_excel)

            else:
                product = product_lookup.get(pn) or products_to_add.get(pn)

                release_year = None
                item_category = None
                original_price = 0
                sale_price = 0

                if product:
                    release_year = product.release_year
                    item_category = product.item_category
                    
                    defaults = product_default_data.get(pn)
                    if defaults:
                        original_price = defaults.get('original_price', 0)
                        sale_price = defaults.get('sale_price', 0)
                        
                    if product_name_from_excel and product.product_name != product_name_from_excel:
                         print(f"Updating product name for existing product {pn}: '{product.product_name}' -> '{product_name_from_excel}'")
                         product.product_name = product_name_from_excel
                         product.product_name_cleaned = clean_string_upper(product_name_from_excel)
                         product.product_name_choseong = get_choseong(product_name_from_excel)

                else:
                    pn_cleaned = clean_string_upper(pn)
                    year_match = re.match(r'^M(2[0-9])', pn_cleaned or '')
                    if year_match:
                        release_year = int(f"20{year_match.group(1)}")
                    
                    final_product_name = product_name_from_excel if product_name_from_excel else f"{pn} (신규)"

                    product = Product(
                        product_number=pn,
                        product_name=final_product_name,
                        product_number_cleaned=pn_cleaned,
                        product_name_cleaned=clean_string_upper(final_product_name),
                        product_name_choseong=get_choseong(final_product_name),
                        release_year=release_year,
                        item_category=item_category
                    )
                    products_to_add[pn] = product
                    product_lookup[pn] = product
                    added_products += 1

                barcode_row_data = {'product_number': pn, 'color': color, 'size': size}
                new_barcode = generate_barcode(barcode_row_data)
                if not new_barcode:
                     print(f"Skipping excel row: Could not generate barcode for {pn}, {color}, {size}")
                     continue

                new_variant_data = {
                    'barcode': new_barcode,
                    'product_number': pn,
                    'color': color,
                    'size': size,
                    'barcode_cleaned': clean_string_upper(new_barcode),
                    'color_cleaned': clean_string_upper(color),
                    'size_cleaned': clean_string_upper(size),
                    'store_stock': 0,
                    'hq_stock': 0,
                    'original_price': original_price,
                    'sale_price': sale_price
                }
                if stock_type == 'store':
                    new_variant_data['store_stock'] = stock_value
                elif stock_type == 'hq':
                    new_variant_data['hq_stock'] = stock_value

                variants_to_add.append(new_variant_data)
                variant_lookup[(pn, color, size)] = True
                added_variants += 1

        print("Loop finished. Committing changes...")
        if products_to_add:
            db.session.add_all(products_to_add.values())
            try:
                db.session.flush()
            except exc.IntegrityError as e:
                 db.session.rollback()
                 flash(f"상품 추가 중 오류 발생 (품번 중복 등): {e}. 재고 업데이트가 취소되었습니다.", 'error')
                 return None, None

        if variants_to_add:
             try:
                 db.session.bulk_insert_mappings(Variant, variants_to_add)
             except exc.IntegrityError as e:
                 db.session.rollback()
                 flash(f"바코드 중복 오류 발생: {e}. 일부 항목이 추가되지 않았을 수 있습니다.", 'error')
                 db.session.commit()
                 return updated_count, added_products

        db.session.commit()
        print("Commit successful.")
        return updated_count, added_variants

    except Exception as e:
        db.session.rollback()
        flash(f"재고 업데이트 중 오류 발생: {e}", 'error')
        print(f"Stock update error: {e}")
        import traceback
        traceback.print_exc()
        return None, None


@app.route('/sync_missing_data', methods=['POST'])
def sync_missing_data():
    updated_variant_count = 0
    updated_product_count = 0
    
    try:
        print("Creating product default data lookup for sync...")
        all_variants = db.session.query(Variant).all()
        all_products = db.session.query(Product).all()

        product_default_lookup = {}
        
        for v in all_variants:
            if v.product_number not in product_default_lookup:
                 product_default_lookup[v.product_number] = {}
            
            if 'original_price' not in product_default_lookup[v.product_number] and v.original_price > 0:
                 product_default_lookup[v.product_number]['original_price'] = v.original_price
            if 'sale_price' not in product_default_lookup[v.product_number] and v.sale_price > 0:
                 product_default_lookup[v.product_number]['sale_price'] = v.sale_price

        for p in all_products:
             if p.product_number not in product_default_lookup:
                 product_default_lookup[p.product_number] = {}

             if 'item_category' not in product_default_lookup[p.product_number] and p.item_category:
                  product_default_lookup[p.product_number]['item_category'] = p.item_category
             if 'release_year' not in product_default_lookup[p.product_number] and p.release_year:
                  product_default_lookup[p.product_number]['release_year'] = p.release_year

        print(f"Default data lookup created with {len(product_default_lookup)} entries.")

        variants_to_update = db.session.query(Variant).filter(
            or_(Variant.original_price == 0, Variant.original_price.is_(None),
                Variant.sale_price == 0, Variant.sale_price.is_(None))
        ).all()
        print(f"Found {len(variants_to_update)} variants to update PRICE.")

        for variant in variants_to_update:
            defaults = product_default_lookup.get(variant.product_number)
            if defaults:
                updated_this_variant = False
                if (variant.original_price is None or variant.original_price == 0) and 'original_price' in defaults:
                    variant.original_price = defaults['original_price']
                    updated_this_variant = True
                if (variant.sale_price is None or variant.sale_price == 0) and 'sale_price' in defaults:
                    variant.sale_price = defaults['sale_price']
                    updated_this_variant = True
                if updated_this_variant:
                    updated_variant_count += 1

        products_to_update = db.session.query(Product).filter(
             or_(Product.item_category.is_(None), Product.item_category == '',
                 Product.release_year.is_(None),
                 Product.product_name_choseong.is_(None)) 
        ).all()
        print(f"Found {len(products_to_update)} products to update INFO.")
        
        year_pattern = re.compile(r'^M(2[0-9])')

        for product in products_to_update:
            defaults = product_default_lookup.get(product.product_number)
            updated_this_product = False
            
            if (not product.item_category) and defaults and 'item_category' in defaults:
                product.item_category = defaults['item_category']
                updated_this_product = True
            
            if product.release_year is None:
                if defaults and 'release_year' in defaults:
                    product.release_year = defaults['release_year']
                    updated_this_product = True
                else:
                    pn_cleaned = product.product_number_cleaned or clean_string_upper(product.product_number)
                    match = year_pattern.match(pn_cleaned)
                    if match:
                        year_short = match.group(1)
                        product.release_year = int(f"20{year_short}")
                        updated_this_product = True
            
            if not product.product_name_choseong:
                product.product_name_choseong = get_choseong(product.product_name)
                updated_this_product = True

            if updated_this_product:
                updated_product_count += 1
        
        if updated_variant_count > 0 or updated_product_count > 0:
            db.session.commit()
            flash(f"동기화 완료: 상품(품목/년도/초성) {updated_product_count}개, SKU(가격) {updated_variant_count}개가 업데이트되었습니다.", 'success')
        else:
            flash("동기화할 데이터가 없거나, 참조할 데이터가 충분하지 않습니다.", 'info')

    except Exception as e:
        db.session.rollback()
        flash(f"동기화 중 오류 발생: {e}", 'error')
        print(f"Sync error: {e}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('stock_management'))


@app.route('/update_store_stock_excel', methods=['POST'])
def update_store_stock_excel():
    file = request.files.get('excel_file')
    updated, added = _process_stock_update_excel(file, request.form, 'store')
    if updated is not None:
        flash(f"매장재고 업데이트 완료: {updated}개 SKU 수정, {added}개 신규 SKU 추가.", 'success')
    return redirect(url_for('stock_management'))

@app.route('/update_hq_stock_excel', methods=['POST'])
def update_hq_stock_excel():
    file = request.files.get('excel_file')
    updated, added = _process_stock_update_excel(file, request.form, 'hq')
    if updated is not None:
        flash(f"본사재고 업데이트 완료: {updated}개 SKU 수정, {added}개 신규 SKU 추가.", 'success')
    return redirect(url_for('stock_management'))


@app.route('/export_stock_check')
def export_stock_check():
    try:
        variants = db.session.query(Variant, Product).join(Product, Variant.product_number == Product.product_number).order_by(Product.product_number, Variant.color, Variant.size).all()
        
        column_order = ['품번', '품명', '컬러', '사이즈', '바코드', '판매가', '매장재고', '실사재고', '과부족', '과부족판매가', '본사재고']
        data_list = []

        for v, p in variants:
            actual_stock_val = v.actual_stock if v.actual_stock is not None else None
            diff_val = (v.store_stock - actual_stock_val) if actual_stock_val is not None else None
            
            data_list.append({
                '품번': v.product_number,
                '품명': p.product_name,
                '컬러': v.color,
                '사이즈': v.size,
                '바코드': v.barcode,
                '판매가': v.sale_price,
                '매장재고': v.store_stock,
                '실사재고': '' if actual_stock_val is None else actual_stock_val,
                '과부족': '' if diff_val is None else diff_val,
                '과부족판매가': '' if diff_val is None else (diff_val * v.sale_price),
                '본사재고': v.hq_stock
            })

        output = _write_dicts_to_excel(data_list, column_order)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='wasabi_stock_check.xlsx')
    except Exception as e: flash(f"엑셀 출력 오류: {e}", 'error'); return redirect(url_for('stock_management'))

def get_sort_key(variant):
    color = variant.color or ''; size_str = str(variant.size).upper().strip()
    custom_order = {'2XS': 'XXS', '2XL': 'XXL', '3XL': 'XXXL'}; size_str = custom_order.get(size_str, size_str)
    order_map = {'XXS': 0, 'XS': 1, 'S': 2, 'M': 3, 'L': 4, 'XL': 5, 'XXL': 6, 'XXXL': 7}
    if size_str.isdigit(): sort_key = (1, int(size_str), '')
    elif size_str in order_map: sort_key = (2, order_map[size_str], '')
    else: sort_key = (3, 0, size_str)
    return (color, sort_key)

@app.route('/product/<product_number>')
def product_detail(product_number):
    product = db.session.get(Product, product_number)
    if product is None: flash("상품 없음.", 'error'); return redirect(url_for('index'))
    product = Product.query.options(joinedload(Product.variants)).get(product_number)

    image_pn = product.product_number.split(' ')[0]; image_url = f"{IMAGE_URL_PREFIX}{image_pn}.jpg"
    variants_list = sorted(product.variants, key=get_sort_key); related_products = []; related_ids = {product_number}

    # (*** [사용자 요청] 수정: 품번 기준 제거, 상품명 기준(요청)으로 변경 ***)
    if product.product_name:
        # 1. 상품명을 공백 기준으로 분리
        search_words = product.product_name.split(' ')
        
        # 2. '남)', '여)' 같은 성별/공용 접두사 제거
        prefixes_to_ignore = ['남)', '여)', '공)', '(남)', '(여)', '(공)', 'M)', 'W)']
        significant_words = [w for w in search_words if w.upper() not in prefixes_to_ignore]

        # 3. 단어가 2개 이상일 경우, 마지막 단어를 제외한 "핵심 이름"을 생성
        # 예: "SPEED ARC MATIS GTX" -> "SPEED ARC MATIS"
        if len(significant_words) > 1:
            # 마지막 단어를 뺀 나머지 단어들로 검색어 생성
            search_term_base = " ".join(significant_words[:-1])
            search_term_cleaned = clean_string_upper(search_term_base) # 예: "SPEEDARCMATIS"
            
            if search_term_cleaned:
                # 4. 이 "핵심 이름"을 포함(LIKE)하는 모든 상품 검색
                # [!!성능 경고!!] 이 쿼리는 DB 인덱스를 사용하지 못해 느릴 수 있습니다.
                name_related = Product.query.options(joinedload(Product.variants)).filter(
                    Product.product_name_cleaned.like(f'%{search_term_cleaned}%'),
                    Product.product_number != product_number
                ).limit(5).all()
                
                for prod in name_related:
                    if prod.product_number not in related_ids:
                        related_products.append(prod)
                        related_ids.add(prod.product_number)

        # 5. (Fallback) 만약 위에서 5개를 못 채웠거나, 단어가 1개였던 경우
        #    가장 마지막 단어(핵심 단어)로도 검색 시도
        if len(related_products) < 5 and significant_words:
            last_word_term = significant_words[-1] # 예: "GTX" 또는 "MATIS"
            last_word_cleaned = clean_string_upper(last_word_term)
            
            # (핵심 단어가 1글자 이상일 때만 검색)
            if last_word_cleaned and len(last_word_cleaned) > 1:
                # [!!성능 경고!!] 이 쿼리도 느릴 수 있습니다.
                last_word_related = Product.query.options(joinedload(Product.variants)).filter(
                    Product.product_name_cleaned.like(f'%{last_word_cleaned}%'),
                    Product.product_number != product_number
                ).limit(5).all()

                for prod in last_word_related:
                    # 5개를 넘기지 않고, 중복이 아닐 경우에만 추가
                    if prod.product_number not in related_ids and len(related_products) < 5:
                        related_products.append(prod)
                        related_ids.add(prod.product_number)

    # (*** [오류 수정] active_page='detail' 변수 추가 ***)
    return render_template(
        'detail.html', 
        product=product, 
        image_url=image_url, 
        variants=variants_list, 
        related_products=related_products[:5],
        active_page='detail'  # <-- [수정] 이 부분이 누락되어 오류가 발생했습니다.
    )

@app.route('/bulk_update_actual_stock', methods=['POST'])
def bulk_update_actual_stock():
    data = request.json; items = data.get('items', [])
    if not items: return jsonify({'status': 'error', 'message': '전송 상품 없음.'}), 400
    try:
        updated = 0; unknown = []
        for item in items:
            original = item.get('barcode', ''); quantity = int(item.get('quantity', 0));

            cleaned_barcode = clean_string_upper(original)
            if not cleaned_barcode or quantity < 0: continue

            stmt = update(Variant).where(Variant.barcode_cleaned == cleaned_barcode).values(actual_stock=quantity)
            result = db.session.execute(stmt); updated += result.rowcount
            if result.rowcount == 0: unknown.append(original)

        db.session.commit()
        msg = f"목록 {len(items)}개 항목 (SKU {updated}개) 실사재고 업데이트 완료."
        if unknown: flash(f"DB에 없는 바코드 {len(unknown)}개: {', '.join(unknown[:5])}...", 'warning')
        flash(msg, 'success'); return jsonify({'status': 'success', 'message': msg})
    except Exception as e: db.session.rollback(); print(f"Bulk update error: {e}"); return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@app.route('/api/fetch_variant', methods=['POST'])
def api_fetch_variant():
    data = request.json; barcode = data.get('barcode', '')
    if not barcode: return jsonify({'status': 'error', 'message': '바코드 없음.'}), 400

    cleaned_barcode = clean_string_upper(barcode)
    if not cleaned_barcode:
        return jsonify({'status': 'error', 'message': f'"{barcode}" 검색 실패.'}), 404

    result = db.session.query(Variant, Product).join(Product, Variant.product_number == Product.product_number).filter(
        Variant.barcode_cleaned == cleaned_barcode
    ).first()

    if result: v, p = result; return jsonify({'status': 'success', 'barcode': v.barcode, 'product_number': p.product_number, 'product_name': p.product_name, 'color': v.color, 'size': v.size, 'store_stock': v.store_stock})
    else: return jsonify({'status': 'error', 'message': f'"{barcode}" 상품 없음.'}), 404

@app.route('/api/search_product_by_prefix', methods=['POST'])
def search_product_by_prefix():
    data = request.json
    barcode_prefix = data.get('prefix', '')

    if not barcode_prefix or len(barcode_prefix) != 11:
        return jsonify({'status': 'error', 'message': '잘못된 바코드 접두사입니다.'}), 400

    search_prefix_cleaned = clean_string_upper(barcode_prefix)

    results = Product.query.filter(
        Product.product_number_cleaned.startswith(search_prefix_cleaned)
    ).all()

    if len(results) == 1:
        return jsonify({'status': 'success', 'product_number': results[0].product_number})
    elif len(results) > 1:
        return jsonify({'status': 'found_many', 'query': barcode_prefix})
    else:
        return jsonify({'status': 'error', 'message': f'"{barcode_prefix}"(으)로 시작하는 품번을 찾을 수 없습니다.'}), 404

@app.route('/update_stock', methods=['POST'])
def update_stock():
    data = request.json; barcode = data.get('barcode'); change = data.get('change')
    if not barcode or change is None: return jsonify({'status': 'error', 'message': '필수 데이터 누락.'}), 400
    try:
        change = int(change); assert change in [1, -1]

        item = db.session.get(Variant, barcode)
        if item is None:
            return jsonify({'status': 'error', 'message': '상품(바코드) 없음.'}), 404

        new_stock = max(0, item.store_stock + change); item.store_stock = new_stock; db.session.commit()
        diff = new_stock - item.actual_stock if item.actual_stock is not None else None
        return jsonify({'status': 'success', 'new_quantity': new_stock, 'barcode': item.barcode, 'new_stock_diff': diff if diff is not None else ''})
    except Exception as e: db.session.rollback(); return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@app.route('/toggle_favorite', methods=['POST'])
def toggle_favorite():
    data = request.json; pn = data.get('product_number')
    if not pn: return jsonify({'status': 'error', 'message': '상품 번호 없음.'}), 400
    try:
        product = db.session.get(Product, pn)
        if product is None: return jsonify({'status': 'error', 'message': '상품 없음.'}), 404
        product.is_favorite = 1 - product.is_favorite; new_status = product.is_favorite
        db.session.commit(); return jsonify({'status': 'success', 'new_favorite_status': new_status})
    except Exception as e: db.session.rollback(); return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@app.route('/update_actual_stock', methods=['POST'])
def update_actual_stock():
    data = request.json; barcode = data.get('barcode'); actual_str = data.get('actual_stock')
    if not barcode: return jsonify({'status': 'error', 'message': '바코드 누락.'}), 400
    try:
        actual = int(actual_str) if actual_str and actual_str.isdigit() else None
        if actual is not None and actual < 0: actual = 0

        item = db.session.get(Variant, barcode)
        if item is None:
            return jsonify({'status': 'error', 'message': '상품(바코드) 없음.'}), 404

        item.actual_stock = actual; db.session.commit()
        diff = item.store_stock - actual if actual is not None else None
        return jsonify({ 'status': 'success', 'barcode': item.barcode, 'new_actual_stock': actual if actual is not None else '', 'new_stock_diff': diff if diff is not None else '' })
    except Exception as e: db.session.rollback(); return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@app.route('/api/update_product_details', methods=['POST'])
def api_update_product_details():
    data = request.json
    pn = data.get('product_number')
    if not pn:
        return jsonify({'status': 'error', 'message': '상품 번호 누락'}), 400

    try:
        product = db.session.get(Product, pn)
        if not product:
            return jsonify({'status': 'error', 'message': '상품을 찾을 수 없음'}), 404

        product.product_name = data.get('product_name', product.product_name)
        product.product_name_cleaned = clean_string_upper(product.product_name)
        product.product_name_choseong = get_choseong(product.product_name)
        try:
            year_val = data.get('release_year')
            product.release_year = int(year_val) if year_val else None
        except (ValueError, TypeError):
            product.release_year = None
        product.item_category = data.get('item_category', product.item_category)

        variants_data = data.get('variants', [])
        barcodes_to_delete = []
        variants_to_add = []
        variants_to_update = {}

        for v_data in variants_data:
            action = v_data.get('action')
            barcode = v_data.get('barcode')

            if action == 'delete' and barcode:
                barcodes_to_delete.append(barcode)
            elif action == 'add':
                variant_row = {
                    'product_number': pn,
                    'color': v_data.get('color'),
                    'size': v_data.get('size'),
                }
                new_barcode = generate_barcode(variant_row)
                if not new_barcode:
                    raise ValueError(f"새 Variant 바코드 생성 실패: {variant_row}")

                variants_to_add.append(Variant(
                    barcode=new_barcode,
                    product_number=pn,
                    color=variant_row['color'],
                    size=variant_row['size'],
                    original_price=int(v_data.get('original_price', 0)),
                    sale_price=int(v_data.get('sale_price', 0)),
                    barcode_cleaned=clean_string_upper(new_barcode),
                    color_cleaned=clean_string_upper(variant_row['color']),
                    size_cleaned=clean_string_upper(variant_row['size'])
                ))
            elif action == 'update' and barcode:
                variants_to_update[barcode] = {
                    'color': v_data.get('color'),
                    'size': v_data.get('size'),
                    'original_price': int(v_data.get('original_price', 0)),
                    'sale_price': int(v_data.get('sale_price', 0)),
                    'color_cleaned': clean_string_upper(v_data.get('color')),
                    'size_cleaned': clean_string_upper(v_data.get('size'))
                }

        if barcodes_to_delete:
             db.session.execute(delete(Variant).where(Variant.barcode.in_(barcodes_to_delete)))

        if variants_to_update:
            existing_variants = Variant.query.filter(Variant.barcode.in_(variants_to_update.keys())).all()
            for variant in existing_variants:
                updates = variants_to_update.get(variant.barcode)
                if updates:
                    variant.color = updates['color']
                    variant.size = updates['size']
                    variant.original_price = updates['original_price']
                    variant.sale_price = updates['sale_price']
                    variant.color_cleaned = updates['color_cleaned']
                    variant.size_cleaned = updates['size_cleaned']

        if variants_to_add:
            db.session.add_all(variants_to_add)

        db.session.commit()
        return jsonify({'status': 'success', 'message': '상품 정보가 업데이트되었습니다.'})

    except ValueError as ve:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'입력 값 오류: {ve}'}), 400
    except exc.IntegrityError as ie:
         db.session.rollback()
         return jsonify({'status': 'error', 'message': f'데이터베이스 오류 (중복 등): {ie}'}), 400
    except Exception as e:
        db.session.rollback()
        print(f"Update product error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500


# (*** /health 엔드포인트는 Render Cron Job 방식이므로 삭제 ***)


@app.cli.command("init-db")
def init_db_command(): init_db()

# (*** 수정: keep_db_awake 함수가 app 컨텍스트를 사용하도록 수정 ***)
def keep_db_awake():
    try:
        # 백그라운드 스레드에서 DB 작업을 수행하려면
        # 반드시 app_context가 필요합니다.
        with app.app_context():
            db.session.execute(text('SELECT 1'))
            print("Neon DB keep-awake (from scheduler).")
    except Exception as e:
        print(f"Keep-awake scheduler error: {e}")

# (*** 수정: 스케줄러를 Gunicorn --preload와 함께 실행되도록 글로벌 범위로 이동 ***)
# Gunicorn 마스터 프로세스에서 이 코드가 실행되어 DB를 깨웁니다.
scheduler = BackgroundScheduler(daemon=True)
scheduler.add_job(keep_db_awake, 'interval', minutes=4) # 4분마다 실행
scheduler.start()
print("APScheduler started (global scope, for Gunicorn --preload).")


if __name__ == '__main__':
    # 이 블록은 'python app.py'로 직접 실행할 때만 동작합니다.
    # Gunicorn 배포 시에는 이 코드가 실행되지 않습니다.
    app.run(debug=True, host='0.0.0.0', port=5000)