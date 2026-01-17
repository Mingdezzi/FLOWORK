import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import io
import re
from collections import Counter
from sqlalchemy import exc, delete
from sqlalchemy.orm import joinedload

from flowork.models import db, Product, Variant, Store, StoreStock
from flowork.utils import generate_barcode, clean_string_upper, get_choseong

def _read_excel_by_letter_to_dicts(file_bytes, col_map, dtype_map):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    col_idx_map = {}
    for key, col_letter in col_map.items():
        if col_letter:
            try:
                col_idx_map[key] = column_index_from_string(col_letter.strip().upper())
            except Exception:
                raise ValueError(f"'{col_letter}'는 유효한 엑셀 열 문자가 아닙니다.")
        else:
            raise ValueError(f"'{key}'에 대한 열 문자가 선택되지 않았습니다.")

    data_rows = []
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        all_none = True
        
        for key, col_idx in col_idx_map.items():
            cell_val = row[col_idx - 1].value
            
            if cell_val is None:
                cell_val = ''
            
            if key in dtype_map and dtype_map[key] == str:
                cell_val = str(cell_val)
                
            if cell_val != '':
                all_none = False
            
            row_data[key] = cell_val
            
        if not all_none:
            data_rows.append(row_data)
            
    return data_rows

def _read_excel_by_header(file_bytes_inner, dtype_map):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes_inner), data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers) if h}
    data_rows = []
    for row in ws.iter_rows(min_row=2):
        row_data = {}
        all_none = True
        for header, col_idx in header_map.items():
            cell_val = row[col_idx].value
            if cell_val is None: cell_val = ''
            if header in dtype_map and dtype_map[header] == str:
                cell_val = str(cell_val)
            if cell_val != '': all_none = False
            row_data[header] = cell_val
        if not all_none: data_rows.append(row_data)
    return data_rows, header_map.keys()


def import_excel_file(file, brand_id):
    if not file or file.filename == '':
        return (False, '파일 선택 안됨.', 'error')

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return (False, '엑셀 파일만 가능.', 'error')

    try:
        file_bytes = file.read()
        
        dtype_map = {'product_number': str, 'color': str, 'size': str, 'product_name': str}
        data, df_columns = _read_excel_by_header(file_bytes, dtype_map)
        
        required_cols = ['product_number', 'product_name', 'color', 'size', 'release_year', 'item_category', 'original_price', 'sale_price', 'is_favorite']
        if not all(col in df_columns for col in required_cols):
             missing = [col for col in required_cols if col not in df_columns]
             return (False, f"엑셀 컬럼명 오류. 누락: {missing}", 'error')

        generated_barcodes = []
        valid_rows = []
        for row in data:
            row['generated_barcode'] = generate_barcode(row)
            if row['generated_barcode']:
                generated_barcodes.append(row['generated_barcode'])
                valid_rows.append(row)
            else:
                print(f"Barcode generation failed for row: {row.get('product_number')}")
        
        if len(valid_rows) != len(data):
            print(f"바코드 생성 실패: {len(data) - len(valid_rows)}개 행 누락.")
        
        duplicates = [item for item, count in Counter(generated_barcodes).items() if count > 1]
        if duplicates:
            return (False, f"바코드 생성 실패: 엑셀 내 중복 발생 ({', '.join(duplicates[:3])}...).", 'error')

        if not valid_rows: 
            return (False, "처리할 유효 데이터 없음.", 'error')

        print(f"Deleting old Product and Variant data for brand_id: {brand_id}...")
        
        products_to_delete = db.session.query(Product.id).filter_by(brand_id=brand_id).all()
        product_ids_to_delete = [p[0] for p in products_to_delete]
        
        if product_ids_to_delete:
            variants_to_delete = db.session.query(Variant.id).filter(Variant.product_id.in_(product_ids_to_delete)).all()
            variant_ids_to_delete = [v[0] for v in variants_to_delete]
            
            if variant_ids_to_delete:
                db.session.execute(delete(StoreStock).where(StoreStock.variant_id.in_(variant_ids_to_delete)))
            
            db.session.execute(delete(Variant).where(Variant.id.in_(variant_ids_to_delete)))
            
        db.session.execute(delete(Product).where(Product.id.in_(product_ids_to_delete)))
        
        print("Old data for this brand cleared.")

        products_data = {}
        products_to_add = []
        variants_data = []
        
        for row in valid_rows:
            pn = row['product_number']
            if not pn: continue

            if pn not in products_data:
                try:
                    year = int(row.get('release_year')) if str(row.get('release_year')).isdigit() else None
                except (ValueError, TypeError):
                    year = None
                
                product_name = str(row.get('product_name', '')) or f"{pn} (신규)"
                
                new_product = Product(
                    brand_id=brand_id,
                    product_number=pn,
                    product_name=product_name,
                    release_year=year,
                    item_category=str(row.get('item_category', '')) or None,
                    is_favorite=int(row.get('is_favorite', 0) or 0),
                    product_number_cleaned=clean_string_upper(pn),
                    product_name_cleaned=clean_string_upper(product_name),
                    product_name_choseong=get_choseong(product_name)
                )
                products_data[pn] = new_product
                products_to_add.append(new_product)

        db.session.add_all(products_to_add)
        db.session.flush()

        for row in valid_rows:
            pn = row['product_number']
            if not pn: continue
            
            product_obj = products_data.get(pn)
            if not product_obj: continue

            variants_data.append({
                'barcode': row['generated_barcode'],
                'product_id': product_obj.id,
                'color': str(row.get('color', '')),
                'size': str(row.get('size', '')),
                'original_price': int(row.get('original_price', 0) or 0),
                'sale_price': int(row.get('sale_price', 0) or 0),
                'barcode_cleaned': clean_string_upper(row['generated_barcode']),
                'color_cleaned': clean_string_upper(row.get('color', '')),
                'size_cleaned': clean_string_upper(row.get('size', ''))
            })

        db.session.bulk_insert_mappings(Variant, variants_data)
        
        db.session.commit()
        return (True, f"성공 ({file.filename}): {len(products_data)}개 상품, {len(variants_data)}개 SKU 임포트. (이 브랜드의 기존 상품/재고 삭제됨)", 'success')
    except exc.IntegrityError as ie:
        db.session.rollback()
        return (False, f"임포트 오류: 바코드 또는 품번이 다른 브랜드와 중복될 수 없습니다. {ie.orig}", 'error')
    except Exception as e:
        db.session.rollback()
        print(f"Import error: {e}")
        import traceback; traceback.print_exc()
        return (False, f"임포트 오류: {e}", 'error')
    

def _write_dicts_to_excel(data_list, column_order):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DB_Export'
    
    ws.append(column_order)
    
    for row_data in data_list:
        row_to_append = [row_data.get(col_name, '') for col_name in column_order]
        ws.append(row_to_append)
        
    for i, col_letter in enumerate(get_column_letter(idx) for idx in range(1, len(column_order) + 1)):
        ws.column_dimensions[col_letter].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_db_to_excel(brand_id):
    try:
        results = db.session.query(
            Product.product_number,
            Product.product_name,
            Variant.color,
            Variant.size,
            Product.release_year,
            Product.item_category,
            Variant.original_price,
            Variant.sale_price,
            Product.is_favorite,
            Variant.barcode
        ).join(Variant, Product.id == Variant.product_id).filter(
            Product.brand_id == brand_id
        ).order_by(
            Product.product_number, Variant.color, Variant.size
        ).all()

        if not results:
            return (None, None, "이 브랜드에 상품/재고 데이터가 없습니다.")

        column_order = [
            'product_number', 'product_name', 'color', 'size',
            'release_year', 'item_category', 'original_price', 'sale_price',
            'is_favorite', 'barcode'
        ]
        
        data_list = []
        for row in results:
            data_list.append({
                'product_number': row.product_number,
                'product_name': row.product_name,
                'color': row.color,
                'size': row.size,
                'release_year': '' if row.release_year is None else row.release_year,
                'item_category': '' if row.item_category is None else row.item_category,
                'original_price': row.original_price,
                'sale_price': row.sale_price,
                'is_favorite': row.is_favorite,
                'barcode': row.barcode
            })

        output = _write_dicts_to_excel(data_list, column_order)
        download_name = 'flowork_product_db_export.xlsx'
        
        return (output, download_name, None)

    except Exception as e:
        print(f"DB Export error: {e}")
        import traceback
        traceback.print_exc()
        return (None, None, f"DB 엑셀 출력 오류: {e}")

def export_stock_check_excel(store_id):
    try:
        store = db.session.get(Store, store_id)
        if not store:
            return (None, None, "매장 정보를 찾을 수 없습니다.")
        
        brand_id = store.brand_id
        
        hq_store = Store.query.filter_by(brand_id=brand_id, is_hq=True).first()
        hq_store_id = hq_store.id if hq_store else None
        
        all_variants_in_brand = db.session.query(Variant).join(Product).filter(
            Product.brand_id == brand_id
        ).options(
            joinedload(Variant.product)
        ).order_by(Product.product_number, Variant.color, Variant.size).all()
        
        if not all_variants_in_brand:
            return (None, None, "이 브랜드에 상품이 없습니다.")

        variant_ids = [v.id for v in all_variants_in_brand]
        
        my_store_stock = db.session.query(StoreStock).filter(
            StoreStock.store_id == store_id,
            StoreStock.variant_id.in_(variant_ids)
        ).all()
        my_stock_map = {s.variant_id: s for s in my_store_stock}
        
        hq_stock_map = {}
        if hq_store_id:
            hq_store_stock = db.session.query(StoreStock).filter(
                StoreStock.store_id == hq_store_id,
                StoreStock.variant_id.in_(variant_ids)
            ).all()
            hq_stock_map = {s.variant_id: s for s in hq_store_stock}

        column_order = ['품번', '품명', '컬러', '사이즈', '바코드', '판매가', '매장재고', '실사재고', '과부족', '과부족판매가', '본사재고']
        data_list = []

        for v in all_variants_in_brand:
            p = v.product
            
            my_stock = my_stock_map.get(v.id)
            hq_stock = hq_stock_map.get(v.id)
            
            store_qty = my_stock.quantity if my_stock else 0
            actual_qty = my_stock.actual_stock if (my_stock and my_stock.actual_stock is not None) else None
            hq_qty = hq_stock.quantity if hq_stock else 0
            
            diff_val = (store_qty - actual_qty) if actual_qty is not None else None
            
            data_list.append({
                '품번': p.product_number,
                '품명': p.product_name,
                '컬러': v.color,
                '사이즈': v.size,
                '바코드': v.barcode,
                '판매가': v.sale_price,
                '매장재고': store_qty,
                '실사재고': '' if actual_qty is None else actual_qty,
                '과부족': '' if diff_val is None else diff_val,
                '과부족판매가': '' if diff_val is None else (diff_val * v.sale_price),
                '본사재고': hq_qty
            })

        output = _write_dicts_to_excel(data_list, column_order)
        download_name = f'flowork_stock_check_{store.store_name}.xlsx'
        
        return (output, download_name, None)
        
    except Exception as e: 
        print(f"Stock Check Export error: {e}")
        import traceback
        traceback.print_exc()
        return (None, None, f"엑셀 출력 오류: {e}")

def _process_stock_update_excel(file, form_data, stock_type, brand_id, store_id_to_update):
    if not file or file.filename == '':
        return (None, None, '파일 선택 안됨.', 'error')

    col_pn_letter = form_data.get('col_product_number')
    col_pname_letter = form_data.get('col_product_name')
    col_color_letter = form_data.get('col_color')
    col_size_letter = form_data.get('col_size')
    col_stock_letter = form_data.get('col_stock')

    if not all([col_pn_letter, col_pname_letter, col_color_letter, col_size_letter, col_stock_letter]):
        return (None, None, '모든 항목(품번, 품명, 컬러, 사이즈, 재고)의 엑셀 열 문자를 선택해야 합니다.', 'error')

    try:
        file_bytes = file.read()
        
        col_map = {
            'product_number': col_pn_letter,
            'product_name': col_pname_letter,
            'color': col_color_letter,
            'size': col_size_letter,
            'stock': col_stock_letter
        }
        dtype_map = {
            'product_number': str, 
            'product_name': str, 
            'color': str, 
            'size': str
        }
        
        data = _read_excel_by_letter_to_dicts(file_bytes, col_map, dtype_map)
        
        print(f"Fetching all products for brand {brand_id}...")
        product_lookup = {
            p.product_number: p for p in 
            db.session.query(Product).filter_by(brand_id=brand_id).all()
        }
        print(f"Product lookup created with {len(product_lookup)} products.")

        print(f"Fetching all variants for brand {brand_id}...")
        variant_lookup = {
            v.barcode: v for v in
            db.session.query(Variant).join(Product).filter(Product.brand_id == brand_id).all()
        }
        print(f"Variant lookup created with {len(variant_lookup)} variants.")

        print(f"Fetching stock for store {store_id_to_update}...")
        stock_lookup = {
            s.variant_id: s for s in
            db.session.query(StoreStock).filter_by(store_id=store_id_to_update).all()
        }
        print(f"Stock lookup created with {len(stock_lookup)} items.")

        updated_count = 0
        added_variants = 0
        added_products = 0
        added_stock_items = 0
        
        variants_to_add = []
        products_to_add = {} 
        stock_to_add = []

        for row in data:
            pn = str(row.get('product_number', '')).strip()
            product_name_from_excel = str(row.get('product_name', '')).strip()
            color = str(row.get('color', '')).strip()
            size = str(row.get('size', '')).strip()
            stock_value_raw = row.get('stock', None)

            try:
                stock_value = int(stock_value_raw) if stock_value_raw is not None and str(stock_value_raw).isdigit() else None
            except (ValueError, TypeError):
                stock_value = None

            if not pn or stock_value is None or stock_value < 0:
                print(f"Skipping excel row: Invalid data (PN: {pn}, Stock: {stock_value_raw})")
                continue
            
            barcode_row_data = {'product_number': pn, 'color': color, 'size': size}
            barcode = generate_barcode(barcode_row_data)
            if not barcode:
                 print(f"Skipping excel row: Could not generate barcode for {pn}, {color}, {size}")
                 continue

            variant = variant_lookup.get(barcode)
            
            if not variant:
                product = product_lookup.get(pn) or products_to_add.get(pn)

                if not product:
                    pn_cleaned = clean_string_upper(pn)
                    year_match = re.match(r'^M(2[0-9])', pn_cleaned or '')
                    release_year = int(f"20{year_match.group(1)}") if year_match else None
                    
                    final_product_name = product_name_from_excel if product_name_from_excel else f"{pn} (신규)"

                    product = Product(
                        brand_id=brand_id,
                        product_number=pn,
                        product_name=final_product_name,
                        product_number_cleaned=pn_cleaned,
                        product_name_cleaned=clean_string_upper(final_product_name),
                        product_name_choseong=get_choseong(final_product_name),
                        release_year=release_year
                    )
                    products_to_add[pn] = product
                    product_lookup[pn] = product
                    db.session.add(product)
                    db.session.flush()
                    added_products += 1
                
                variant = Variant(
                    barcode=barcode,
                    product_id=product.id,
                    color=color,
                    size=size,
                    barcode_cleaned=clean_string_upper(barcode),
                    color_cleaned=clean_string_upper(color),
                    size_cleaned=clean_string_upper(size)
                )
                variants_to_add.append(variant)
                variant_lookup[barcode] = variant
                db.session.add(variant)
                db.session.flush()
                added_variants += 1

            stock_item = stock_lookup.get(variant.id)

            if stock_item:
                stock_item.quantity = stock_value
                updated_count += 1
            else:
                stock_item = StoreStock(
                    store_id=store_id_to_update,
                    variant_id=variant.id,
                    quantity=stock_value
                )
                stock_to_add.append(stock_item)
                stock_lookup[variant.id] = stock_item 
                added_stock_items += 1

        if stock_to_add:
            db.session.add_all(stock_to_add)

        db.session.commit()
        print("Commit successful.")
        
        stock_type_kor = "매장" if stock_type == "store" else "본사"
        message = f"{stock_type_kor}재고 업데이트 완료: {updated_count}개 SKU 수정, {added_stock_items}개 신규 SKU 재고 등록. (신규 상품 {added_products}개, 신규 옵션 {added_variants}개)"
        return (updated_count, added_stock_items, message, 'success')

    except exc.IntegrityError as ie:
        db.session.rollback()
        return (None, None, f"재고 업데이트 중 오류 발생 (바코드 중복 등): {ie.orig}", 'error')
    except Exception as e:
        db.session.rollback()
        print(f"Stock update error: {e}")
        import traceback
        traceback.print_exc()
        return (None, None, f"재고 업데이트 중 오류 발생: {e}", 'error')