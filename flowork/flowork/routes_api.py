import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from flask import Blueprint, request, redirect, url_for, flash, jsonify, send_file, abort
import io
import traceback
from flask_login import login_required, current_user
from datetime import datetime # (신규) datetime 임포트

# (수정) Order, OrderProcessing, Staff, ScheduleEvent 임포트
from flowork.models import db, Product, Variant, Order, OrderProcessing, Announcement, Store, Setting, Brand, StoreStock, Staff, ScheduleEvent
from flowork.utils import clean_string_upper, get_choseong, generate_barcode
from sqlalchemy import or_, update, exc, delete, func
from sqlalchemy.orm import joinedload, selectinload

from flowork.services_excel import (
    import_excel_file,
    export_db_to_excel,
    export_stock_check_excel,
    _process_stock_update_excel 
)
from flowork.services_db import sync_missing_data_in_db

api_bp = Blueprint('api', __name__)

from functools import wraps
def admin_required(f):
    @wraps(f)
    @login_required 
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            abort(403, description="이 작업을 수행할 관리자 권한이 없습니다.")
        return f(*args, **kwargs)
    return decorated_function

# (수정) session 인자 제거 및 Race Condition 방어 로직 추가
def _get_or_create_store_stock(variant_id, store_id):
    # 1. 먼저 해당 재고가 있는지 확인
    stock = db.session.query(StoreStock).filter_by(
        variant_id=variant_id,
        store_id=store_id
    ).first()
    
    if stock:
        return stock

    # 2. 없으면 새로 생성 시도
    try:
        stock = StoreStock(
            variant_id=variant_id,
            store_id=store_id,
            quantity=0,
            actual_stock=None
        )
        db.session.add(stock)
        db.session.commit() # (수정) 추가 후 즉시 commit
        return stock
    except exc.IntegrityError:
        # 3. 만약 다른 요청이 동시에 생성해서 commit에 실패했다면,
        # 롤백하고 다시 조회하여 기존에 생성된 객체를 반환
        db.session.rollback()
        stock = db.session.query(StoreStock).filter_by(
            variant_id=variant_id,
            store_id=store_id
        ).first()
        return stock

def _parse_iso_date_string(date_str):
    """ (신규) ISO 날짜 문자열(YYYY-MM-DD)을 datetime.date 객체로 변환 """
    if not date_str:
        return None
    try:
        # YYYY-MM-DD 형식으로 가정
        return datetime.strptime(date_str.split('T')[0], '%Y-%m-%d').date()
    except ValueError:
        print(f"Warning: Could not parse date string {date_str}")
        return None

@api_bp.route('/api/setting/brand_name', methods=['POST'])
@admin_required
def update_brand_name():
    data = request.json
    brand_name = data.get('brand_name', '').strip()
    
    if not brand_name:
        return jsonify({'status': 'error', 'message': '브랜드 이름이 비어있습니다.'}), 400
        
    try:
        brand = db.session.get(Brand, current_user.store.brand_id)
        if not brand:
            return jsonify({'status': 'error', 'message': '브랜드를 찾을 수 없습니다.'}), 404
            
        brand.brand_name = brand_name
        
        brand_name_setting = Setting.query.filter_by(
            store_id=current_user.store_id, 
            key='BRAND_NAME'
        ).first()
        if not brand_name_setting:
            brand_name_setting = Setting(store_id=current_user.store_id, key='BRAND_NAME')
            db.session.add(brand_name_setting)
        brand_name_setting.value = brand_name
        
        db.session.commit()
        
        return jsonify({
            'status': 'success', 
            'message': f"브랜드 이름이 '{brand_name}'(으)로 저장되었습니다.",
            'brand_name': brand_name
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error updating brand name: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500


@api_bp.route('/api/stores', methods=['GET'])
@login_required
def get_stores():
    try:
        stores = Store.query.filter_by(
            brand_id=current_user.store.brand_id 
        ).order_by(Store.store_name).all()
        
        return jsonify({
            'status': 'success',
            'stores': [{
                'store_id': s.id, 
                'store_name': s.store_name,
                'store_phone': s.phone_number or ''
            } for s in stores]
        })
    except Exception as e:
        print(f"Error getting stores: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/stores', methods=['POST'])
@admin_required
def add_store():
    data = request.json
    name = data.get('store_name', '').strip()
    phone = data.get('store_phone', '').strip()
    is_hq = data.get('is_hq', False)

    if not name:
        return jsonify({'status': 'error', 'message': '매장 이름은 필수입니다.'}), 400
    
    try:
        existing_name = Store.query.filter(
            Store.brand_id == current_user.store.brand_id, 
            func.lower(Store.store_name) == func.lower(name)
        ).first()
        if existing_name:
            return jsonify({'status': 'error', 'message': f"매장 이름 '{name}'(이)가 이미 존재합니다."}), 409

        new_store = Store(
            brand_id=current_user.store.brand_id, 
            store_name=name,
            phone_number=phone,
            is_hq=is_hq
        )
        db.session.add(new_store)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f"'{name}'(이)가 추가되었습니다.",
            'store': {
                'store_id': new_store.id, 
                'store_name': new_store.store_name,
                'store_phone': new_store.phone_number or '',
                'is_hq': new_store.is_hq
            }
        }), 201 
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding store: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/stores/<int:store_id>', methods=['POST'])
@admin_required
def update_store(store_id):
    data = request.json
    name = data.get('store_name', '').strip()
    phone = data.get('store_phone', '').strip()
    is_hq = data.get('is_hq', False)

    if not name:
         return jsonify({'status': 'error', 'message': '매장 이름은 필수입니다.'}), 400

    try:
        store = Store.query.filter_by(
            id=store_id, 
            brand_id=current_user.store.brand_id
        ).first()
        
        if not store:
            return jsonify({'status': 'error', 'message': '수정할 매장을 찾을 수 없습니다.'}), 404

        existing_name = Store.query.filter(
            Store.brand_id == current_user.store.brand_id, 
            func.lower(Store.store_name) == func.lower(name),
            Store.id != store_id
        ).first()
        if existing_name:
            return jsonify({'status': 'error', 'message': f"매장 이름 '{name}'(이)가 이미 존재합니다."}), 409

        store.store_name = name
        store.phone_number = phone
        store.is_hq = is_hq
        db.session.commit()
        message = f"'{name}' 정보가 수정되었습니다."

        return jsonify({
            'status': 'success',
            'message': message,
            'store': {
                'store_id': store.id, 
                'store_name': store.store_name,
                'store_phone': store.phone_number or '',
                'is_hq': store.is_hq
            }
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error updating store: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500


@api_bp.route('/api/stores/<int:store_id>', methods=['DELETE'])
@admin_required
def delete_store(store_id):
    try:
        if current_user.store_id == store_id:
            return jsonify({'status': 'error', 'message': "현재 로그인된 매장은 삭제할 수 없습니다."}), 403

        store = Store.query.filter_by(
            id=store_id, 
            brand_id=current_user.store.brand_id
        ).first()
        
        if not store:
            return jsonify({'status': 'error', 'message': '삭제할 매장을 찾을 수 없습니다.'}), 404
            
        name = store.store_name
        db.session.delete(store)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f"'{name}'(이)가 삭제되었습니다."
        })
        
    except exc.IntegrityError:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f"'{name}'(은)는 현재 사용자 또는 주문/재고 내역에서 사용 중이므로 삭제할 수 없습니다."}), 409
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting store: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

# --- (신규) 직원(Staff) API ---
@api_bp.route('/api/staff', methods=['POST'])
@admin_required
def add_staff():
    data = request.json
    name = data.get('name', '').strip()
    position = data.get('position', '').strip()
    contact = data.get('contact', '').strip()

    if not name:
        return jsonify({'status': 'error', 'message': '직원 이름은 필수입니다.'}), 400
    
    try:
        new_staff = Staff(
            store_id=current_user.store_id,
            name=name,
            position=position or None,
            contact=contact or None,
            is_active=True
        )
        db.session.add(new_staff)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f"직원 '{name}'(이)가 추가되었습니다.",
            'staff': {
                'id': new_staff.id, 
                'name': new_staff.name,
                'position': new_staff.position or '',
                'contact': new_staff.contact or ''
            }
        }), 201 
        
    except Exception as e:
        db.session.rollback()
        print(f"Error adding staff: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/staff/<int:staff_id>', methods=['POST'])
@admin_required
def update_staff(staff_id):
    data = request.json
    name = data.get('name', '').strip()
    position = data.get('position', '').strip()
    contact = data.get('contact', '').strip()

    if not name:
         return jsonify({'status': 'error', 'message': '직원 이름은 필수입니다.'}), 400

    try:
        staff = Staff.query.filter_by(
            id=staff_id, 
            store_id=current_user.store_id
        ).first()
        
        if not staff:
            return jsonify({'status': 'error', 'message': '수정할 직원을 찾을 수 없습니다.'}), 404

        staff.name = name
        staff.position = position or None
        staff.contact = contact or None
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f"'{name}' 직원 정보가 수정되었습니다.",
            'staff': {
                'id': staff.id, 
                'name': staff.name,
                'position': staff.position or '',
                'contact': staff.contact or ''
            }
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error updating staff: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/staff/<int:staff_id>', methods=['DELETE'])
@admin_required
def delete_staff(staff_id):
    try:
        staff = Staff.query.filter_by(
            id=staff_id, 
            store_id=current_user.store_id
        ).first()
        
        if not staff:
            return jsonify({'status': 'error', 'message': '삭제할 직원을 찾을 수 없습니다.'}), 404
            
        name = staff.name
        
        # (수정) 실제 삭제 대신 is_active = False (휴가 기록 등 보존)
        staff.is_active = False 
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f"'{name}' 직원이 (비활성) 삭제 처리되었습니다."
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting staff: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500
# --- 직원 API 끝 ---


@api_bp.route('/import_excel', methods=['POST'])
@admin_required
def import_excel():
    file = request.files.get('excel_file')
    success, message, category = import_excel_file(file, current_user.store.brand_id)
    flash(message, category)
    return redirect(url_for('ui.setting_page'))

@api_bp.route('/export_db_excel')
@login_required
def export_db_excel():
    output, download_name, error_message = export_db_to_excel(current_user.store.brand_id)
    if error_message:
        flash(error_message, 'warning')
        return redirect(url_for('ui.setting_page'))
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@api_bp.route('/sync_missing_data', methods=['POST'])
@login_required
def sync_missing_data():
    success, message, category = sync_missing_data_in_db(current_user.store.brand_id)
    flash(message, category)
    return redirect(url_for('ui.stock_management'))

@api_bp.route('/update_store_stock_excel', methods=['POST'])
@login_required
def update_store_stock_excel():
    file = request.files.get('excel_file')
    updated, added, message, category = _process_stock_update_excel(
        file, request.form, 'store', 
        current_user.store.brand_id, 
        current_user.store_id
    )
    flash(message, category)
    return redirect(url_for('ui.stock_management'))

@api_bp.route('/update_hq_stock_excel', methods=['POST'])
@login_required
def update_hq_stock_excel():
    file = request.files.get('excel_file')
    hq_store = Store.query.filter_by(brand_id=current_user.store.brand_id, is_hq=True).first()
    if not hq_store:
        flash("본사(HQ) 매장이 설정되지 않아 본사재고를 업데이트할 수 없습니다.", "error")
        return redirect(url_for('ui.stock_management'))
        
    updated, added, message, category = _process_stock_update_excel(
        file, request.form, 'hq', 
        current_user.store.brand_id, 
        hq_store.id
    )
    flash(message, category)
    return redirect(url_for('ui.stock_management'))

@api_bp.route('/export_stock_check')
@login_required
def export_stock_check():
    output, download_name, error_message = export_stock_check_excel(current_user.store_id)
    if error_message:
        flash(error_message, 'error')
        return redirect(url_for('ui.stock_management'))

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@api_bp.route('/api/live_search', methods=['POST'])
@login_required
def live_search():
    data = request.json
    query_param = data.get('query', '')
    category_param = data.get('category', '전체')

    # (수정) selectinload(Product.variants) 추가 (N+1 해결)
    base_query = Product.query.options(selectinload(Product.variants)).filter(
        Product.brand_id == current_user.store.brand_id
    )
    showing_favorites = False

    is_searching = bool(query_param) or (category_param and category_param != '전체')

    if is_searching:
        if query_param:
            search_term_cleaned = clean_string_upper(query_param)
            search_like = f'%{search_term_cleaned}%'
            base_query = base_query.filter(
                or_(
                    Product.product_number_cleaned.like(search_like),
                    Product.product_name_cleaned.like(search_like),
                    Product.product_name_choseong.like(search_like)
                )
            )

        if category_param and category_param != '전체':
            base_query = base_query.filter(Product.item_category == category_param)

        products = base_query.order_by(Product.release_year.desc(), Product.product_name).all()
    else:
        showing_favorites = True
        products = base_query.filter(Product.is_favorite == 1).order_by(Product.item_category, Product.product_name).all()

    results_list = []
    for product in products:
        image_pn = product.product_number.split(' ')[0]

        colors = ""
        sale_price_f = "가격정보없음"
        original_price_f = 0
        discount_f = "-"

        # (수정) .all() 제거 (selectinload로 이미 로드됨)
        product_variants = product.variants 

        if product_variants:
            colors_list = sorted(list(set(v.color for v in product_variants if v.color)))
            colors = ", ".join(colors_list)
            first_variant = product_variants[0]
            sale_price_f = f"{first_variant.sale_price:,d}원"
            original_price_f = first_variant.original_price
            if original_price_f and original_price_f > 0:
                discount_f = f"{int((1 - (first_variant.sale_price / original_price_f)) * 100)}%"
            else:
                discount_f = "0%"

        results_list.append({
            "product_id": product.id,
            "product_number": product.product_number,
            "product_name": product.product_name,
            "image_pn": image_pn,
            "colors": colors,
            "sale_price": sale_price_f,
            "original_price": original_price_f,
            "discount": discount_f
        })

    return jsonify({
        "status": "success",
        "products": results_list,
        "showing_favorites": showing_favorites,
        "selected_category": category_param
    })

@api_bp.route('/reset_actual_stock', methods=['POST'])
@login_required
def reset_actual_stock():
    try: 
        store_stock_ids_query = db.session.query(StoreStock.id).filter_by(store_id=current_user.store_id)
        
        stmt = update(StoreStock).where(
            StoreStock.id.in_(store_stock_ids_query)
        ).values(actual_stock=None)
        
        result = db.session.execute(stmt)
        db.session.commit()
        flash(f'실사재고 {result.rowcount}건 초기화 완료.', 'success')
    except Exception as e: 
        db.session.rollback()
        flash(f'초기화 오류: {e}', 'error')
    return redirect(url_for('ui.check_page'))

@api_bp.route('/reset_database_completely', methods=['POST'])
@admin_required
def reset_database_completely():
    try:
        print("Deleting Product/Variant/StoreStock/Order tables...")
        engine = db.get_engine(bind=None)
        
        # (수정) Product에 의존하는 Order, OrderProcessing 테이블을 먼저 삭제 리스트에 추가
        tables_to_drop = [
            OrderProcessing.__table__, 
            Order.__table__,
            StoreStock.__table__, 
            Variant.__table__, 
            Product.__table__
        ]
        
        db.Model.metadata.drop_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        db.Model.metadata.create_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        
        # (수정) Flash 메시지 변경
        flash('데이터베이스 초기화 완료. (상품/재고/주문 데이터 삭제됨. 계정/공지 내역 보존)', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'DB 전체 초기화 오류: {e}', 'error')
        print(f"DB Reset Error: {e}")
        traceback.print_exc()
    return redirect(url_for('ui.setting_page'))

@api_bp.route('/api/analyze_excel', methods=['POST'])
@login_required
def analyze_excel():
    if 'excel_file' not in request.files:
        return jsonify({'status': 'error', 'message': '파일이 없습니다.'}), 400
    
    file = request.files.get('excel_file')
    if file.filename == '':
        return jsonify({'status': 'error', 'message': '파일이 선택되지 않았습니다.'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'status': 'error', 'message': '엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.'}), 400

    try:
        file_bytes = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        max_col_index = ws.max_column
        if max_col_index > 26: max_col_index = 26 
        column_letters = [get_column_letter(i) for i in range(1, max_col_index + 1)]
        
        preview_data = {}
        max_row_preview = min(6, ws.max_row + 1) 
        
        if max_row_preview <= 1:
             return jsonify({'status': 'error', 'message': '파일에 데이터가 없습니다.'}), 400

        for col_letter in column_letters:
            col_data = []
            col_index = column_index_from_string(col_letter)
            for i in range(1, max_row_preview):
                cell_val = ws.cell(row=i, column=col_index).value
                col_data.append(str(cell_val) if cell_val is not None else '')
            preview_data[col_letter] = col_data
            
        return jsonify({
            'status': 'success',
            'column_letters': column_letters,
            'preview_data': preview_data
        })
        
    except Exception as e:
        print(f"Excel analyze error: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'엑셀 파일 분석 중 오류 발생: {e}'}), 500

@api_bp.route('/bulk_update_actual_stock', methods=['POST'])
@login_required
def bulk_update_actual_stock():
    data = request.json
    items = data.get('items', [])
    if not items: 
        return jsonify({'status': 'error', 'message': '전송 상품 없음.'}), 400
    
    try:
        updated = 0
        unknown = []
        barcode_map = {clean_string_upper(item.get('barcode', '')): int(item.get('quantity', 0)) for item in items if item.get('barcode')}
        
        if not barcode_map:
            return jsonify({'status': 'error', 'message': '유효한 바코드 없음.'}), 400

        variants = db.session.query(Variant).join(Product).filter(
            Product.brand_id == current_user.store.brand_id,
            Variant.barcode_cleaned.in_(barcode_map.keys())
        ).all()
        
        variant_id_map = {v.barcode_cleaned: v.id for v in variants}
        found_barcodes = set(variant_id_map.keys())
        unknown = [b for b in barcode_map.keys() if b not in found_barcodes]
        
        if not variant_id_map:
            return jsonify({'status': 'error', 'message': 'DB에 일치하는 상품이 없습니다.'}), 404

        existing_stock = db.session.query(StoreStock).filter(
            StoreStock.store_id == current_user.store_id,
            StoreStock.variant_id.in_(variant_id_map.values())
        ).all()
        
        stock_map = {s.variant_id: s for s in existing_stock}
        
        new_stock_entries = []
        for barcode_cleaned, variant_id in variant_id_map.items():
            new_actual_qty = barcode_map[barcode_cleaned]
            
            if variant_id in stock_map:
                stock_map[variant_id].actual_stock = new_actual_qty
                updated += 1
            else:
                new_stock = StoreStock(
                    store_id=current_user.store_id,
                    variant_id=variant_id,
                    quantity=0,
                    actual_stock=new_actual_qty
                )
                new_stock_entries.append(new_stock)
                updated += 1

        if new_stock_entries:
            db.session.add_all(new_stock_entries)
            
        db.session.commit()
        msg = f"목록 {len(items)}개 항목 (SKU {updated}개) 실사재고 업데이트 완료."
        if unknown: 
            flash(f"DB에 없는 바코드 {len(unknown)}개: {', '.join(unknown[:5])}...", 'warning')
        flash(msg, 'success')
        return jsonify({'status': 'success', 'message': msg})
    except Exception as e: 
        db.session.rollback()
        print(f"Bulk update error: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/fetch_variant', methods=['POST'])
@login_required
def api_fetch_variant():
    data = request.json
    barcode = data.get('barcode', '')
    if not barcode: 
        return jsonify({'status': 'error', 'message': '바코드 없음.'}), 400

    cleaned_barcode = clean_string_upper(barcode)
    if not cleaned_barcode:
        return jsonify({'status': 'error', 'message': f'"{barcode}" 검색 실패.'}), 404

    result = db.session.query(Variant, Product).join(Product, Variant.product_id == Product.id).filter(
        Variant.barcode_cleaned == cleaned_barcode,
        Product.brand_id == current_user.store.brand_id
    ).first()

    if result: 
        v, p = result
        
        stock = db.session.query(StoreStock).filter_by(
            variant_id=v.id,
            store_id=current_user.store_id
        ).first()
        
        current_stock_qty = stock.quantity if stock else 0
        
        return jsonify({
            'status': 'success', 
            'barcode': v.barcode, 
            'product_number': p.product_number, 
            'product_name': p.product_name, 
            'color': v.color, 
            'size': v.size, 
            'store_stock': current_stock_qty
        })
    else: 
        return jsonify({'status': 'error', 'message': f'"{barcode}" 상품 없음.'}), 404

@api_bp.route('/api/search_product_by_prefix', methods=['POST'])
@login_required
def search_product_by_prefix():
    data = request.json
    barcode_prefix = data.get('prefix', '')

    if not barcode_prefix or len(barcode_prefix) != 11:
        return jsonify({'status': 'error', 'message': '잘못된 바코드 접두사입니다.'}), 400

    search_prefix_cleaned = clean_string_upper(barcode_prefix)

    results = Product.query.filter(
        Product.brand_id == current_user.store.brand_id,
        Product.product_number_cleaned.startswith(search_prefix_cleaned)
    ).all()

    if len(results) == 1:
        return jsonify({'status': 'success', 'product_number': results[0].product_number})
    elif len(results) > 1:
        return jsonify({'status': 'found_many', 'query': barcode_prefix})
    else:
        return jsonify({'status': 'error', 'message': f'"{barcode_prefix}"(으)로 시작하는 품번을 찾을 수 없습니다.'}), 404

@api_bp.route('/update_stock', methods=['POST'])
@login_required
def update_stock():
    data = request.json
    barcode = data.get('barcode')
    change = data.get('change')
    if not barcode or change is None: 
        return jsonify({'status': 'error', 'message': '필수 데이터 누락.'}), 400
    try:
        change = int(change)
        assert change in [1, -1]

        cleaned_barcode = clean_string_upper(barcode)
        
        variant = db.session.query(Variant).join(Product).filter(
            Variant.barcode_cleaned == cleaned_barcode,
            Product.brand_id == current_user.store.brand_id
        ).first()
        
        if variant is None:
            return jsonify({'status': 'error', 'message': '상품(바코드) 없음.'}), 404
        
        # (수정) session 인자 제거 및 commit 분리
        stock = _get_or_create_store_stock(variant.id, current_user.store_id)
        
        # _get_or_create_store_stock이 commit을 할 수 있으므로,
        # 재고 수량 변경은 별도의 트랜잭션으로 처리
        new_stock = max(0, stock.quantity + change)
        stock.quantity = new_stock
        db.session.commit()
        
        diff = new_stock - stock.actual_stock if stock.actual_stock is not None else None
        return jsonify({
            'status': 'success', 
            'new_quantity': new_stock, 
            'barcode': barcode, 
            'new_stock_diff': diff if diff is not None else ''
        })
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/toggle_favorite', methods=['POST'])
@login_required
def toggle_favorite():
    data = request.json
    product_id = data.get('product_id')
    if not product_id: 
        return jsonify({'status': 'error', 'message': '상품 ID 없음.'}), 400
    try:
        product = Product.query.filter_by(
            id=product_id,
            brand_id=current_user.store.brand_id
        ).first()
        
        if product is None: 
            return jsonify({'status': 'error', 'message': '상품 없음.'}), 404
        
        product.is_favorite = 1 - product.is_favorite
        new_status = product.is_favorite
        db.session.commit()
        return jsonify({'status': 'success', 'new_favorite_status': new_status})
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/update_actual_stock', methods=['POST'])
@login_required
def update_actual_stock():
    data = request.json
    barcode = data.get('barcode')
    actual_str = data.get('actual_stock')
    if not barcode: 
        return jsonify({'status': 'error', 'message': '바코드 누락.'}), 400
    try:
        actual = int(actual_str) if actual_str and actual_str.isdigit() else None
        if actual is not None and actual < 0: 
            actual = 0

        cleaned_barcode = clean_string_upper(barcode)
        
        variant = db.session.query(Variant).join(Product).filter(
            Variant.barcode_cleaned == cleaned_barcode,
            Product.brand_id == current_user.store.brand_id
        ).first()

        if variant is None:
            return jsonify({'status': 'error', 'message': '상품(바코드) 없음.'}), 404

        # (수정) session 인자 제거 및 commit 분리
        stock = _get_or_create_store_stock(variant.id, current_user.store_id)
        
        stock.actual_stock = actual
        db.session.commit()
        
        diff = stock.quantity - actual if actual is not None else None
        return jsonify({ 
            'status': 'success', 
            'barcode': barcode, 
            'new_actual_stock': actual if actual is not None else '', 
            'new_stock_diff': diff if diff is not None else '' 
        })
    except Exception as e: 
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/update_product_details', methods=['POST'])
@admin_required
def api_update_product_details():
    data = request.json
    product_id = data.get('product_id')
    if not product_id:
        return jsonify({'status': 'error', 'message': '상품 ID 누락'}), 400

    try:
        product = Product.query.filter_by(
            id=product_id,
            brand_id=current_user.store.brand_id
        ).first()
        
        if not product:
            return jsonify({'status': 'error', 'message': '상품을 찾을 수 없음'}), 404

        product.product_name = data.get('product_name', product.product_name)
        product.product_name_cleaned = clean_string_upper(product.product_name)
        product.product_name_choseong = get_choseong(product.product_name)
        try:
            year_val = data.get('release_year')
            product.release_year = int(year_val) if year_val else None
        except (ValueError, TypeError):
            product.release_year = None
        product.item_category = data.get('item_category', product.item_category)

        variants_data = data.get('variants', [])
        variant_ids_to_delete = []
        variants_to_add = []
        variants_to_update = {}

        for v_data in variants_data:
            action = v_data.get('action')
            variant_id = v_data.get('variant_id')

            if action == 'delete' and variant_id:
                variant_ids_to_delete.append(variant_id)
            elif action == 'add':
                variant_row = {
                    'product_number': product.product_number,
                    'color': v_data.get('color'),
                    'size': v_data.get('size'),
                }
                new_barcode = generate_barcode(variant_row)
                if not new_barcode:
                    raise ValueError(f"새 Variant 바코드 생성 실패: {variant_row}")
                
                existing_barcode = Variant.query.filter_by(barcode_cleaned=clean_string_upper(new_barcode)).first()
                if existing_barcode:
                    raise exc.IntegrityError(f"바코드 중복: {new_barcode}", params=None, orig=None)

                variants_to_add.append(Variant(
                    barcode=new_barcode,
                    product_id=product.id,
                    color=variant_row['color'],
                    size=variant_row['size'],
                    original_price=int(v_data.get('original_price', 0)),
                    sale_price=int(v_data.get('sale_price', 0)),
                    barcode_cleaned=clean_string_upper(new_barcode),
                    color_cleaned=clean_string_upper(variant_row['color']),
                    size_cleaned=clean_string_upper(variant_row['size'])
                ))
            elif action == 'update' and variant_id:
                variants_to_update[variant_id] = {
                    'color': v_data.get('color'),
                    'size': v_data.get('size'),
                    'original_price': int(v_data.get('original_price', 0)),
                    'sale_price': int(v_data.get('sale_price', 0)),
                    'color_cleaned': clean_string_upper(v_data.get('color')),
                    'size_cleaned': clean_string_upper(v_data.get('size'))
                }

        if variant_ids_to_delete:
             db.session.execute(delete(StoreStock).where(
                 StoreStock.variant_id.in_(variant_ids_to_delete)
             ))
             db.session.execute(delete(Variant).where(
                 Variant.id.in_(variant_ids_to_delete),
                 Variant.product_id == product.id 
             ))

        if variants_to_update:
            existing_variants = Variant.query.filter(
                Variant.id.in_(variants_to_update.keys()),
                Variant.product_id == product.id
            ).all()
            for variant in existing_variants:
                updates = variants_to_update.get(variant.id)
                if updates:
                    variant.color = updates['color']
                    variant.size = updates['size']
                    variant.original_price = updates['original_price']
                    variant.sale_price = updates['sale_price']
                    variant.color_cleaned = updates['color_cleaned']
                    variant.size_cleaned = updates['size_cleaned']

        if variants_to_add:
            db.session.add_all(variants_to_add)

        db.session.commit()
        return jsonify({'status': 'success', 'message': '상품 정보가 업데이트되었습니다.'})

    except ValueError as ve:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'입력 값 오류: {ve}'}), 400
    except exc.IntegrityError as ie:
         db.session.rollback()
         return jsonify({'status': 'error', 'message': f'데이터베이스 오류 (바코드 중복 등): {ie.orig}'}), 400
    except Exception as e:
        db.session.rollback()
        print(f"Update product error: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/find_product_details', methods=['POST'])
@login_required
def api_find_product_details():
    data = request.json
    pn_query = data.get('product_number', '')
    if not pn_query:
        return jsonify({'status': 'error', 'message': '품번 없음.'}), 400
    
    search_term_cleaned = clean_string_upper(pn_query)
    search_like = f"%{search_term_cleaned}%"
    
    product = Product.query.options(
        selectinload(Product.variants)
    ).filter(
        Product.brand_id == current_user.store.brand_id,
        Product.product_number_cleaned.like(search_like)
    ).first()

    if product:
        # (수정) .all() 제거
        variants = sorted(product.variants, key=lambda v: (v.color, v.size)) 
        
        colors = []
        if variants:
            colors_seen = set()
            for v in variants:
                if v.color not in colors_seen:
                    colors.append(v.color)
                    colors_seen.add(v.color)
                    
        sizes = []
        if variants:
            sizes_seen = set()
            for v in variants:
                 if v.size not in sizes_seen:
                    sizes.append(v.size)
                    sizes_seen.add(v.size)
        
        return jsonify({
            'status': 'success',
            'product_name': product.product_name,
            'product_number': product.product_number,
            'colors': colors,
            'sizes': sizes
        })
    else:
        return jsonify({
            'status': 'error',
            'message': f"'{pn_query}'(으)로 시작하는 상품을 찾을 수 없습니다."
        }), 404

@api_bp.route('/api/order_product_search', methods=['POST'])
@login_required
def api_order_product_search():
    data = request.json
    query = data.get('query', '')
    if not query:
        return jsonify({'status': 'error', 'message': '검색어 없음.'}), 400
    
    search_term_cleaned = clean_string_upper(query)
    search_like = f"%{search_term_cleaned}%"
    
    products = Product.query.filter(
        Product.brand_id == current_user.store.brand_id,
        or_(
            Product.product_number_cleaned.like(search_like),
            Product.product_name_cleaned.like(search_like),
            Product.product_name_choseong.like(search_like)
        )
    ).order_by(Product.product_name).limit(20).all()

    if products:
        results = [{
            'product_id': p.id,
            'product_number': p.product_number,
            'product_name': p.product_name
        } for p in products]
        return jsonify({'status': 'success', 'products': results})
    else:
        return jsonify({'status': 'error', 'message': f"'{query}'(으)로 검색된 상품이 없습니다."}), 404

@api_bp.route('/api/update_order_status', methods=['POST'])
@login_required
def api_update_order_status():
    data = request.json
    order_id = data.get('order_id')
    new_status = data.get('new_status')

    if not order_id or not new_status:
        return jsonify({'status': 'error', 'message': '필수 정보 누락'}), 400
    
    try:
        order = Order.query.filter_by(
            id=order_id, 
            store_id=current_user.store_id
        ).first()
        
        if not order:
            return jsonify({'status': 'error', 'message': '주문을 찾을 수 없거나 권한이 없습니다.'}), 404
        
        order.order_status = new_status
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': f'주문(ID: {order_id}) 상태가 {new_status}(으)로 변경되었습니다.'})

    except Exception as e:
        db.session.rollback()
        print(f"Error updating order status: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500


@api_bp.route('/api/reset-orders-db', methods=['POST'])
@admin_required
def reset_orders_db():
    try:
        engine = db.get_engine(bind=None)
        if engine is None:
            raise Exception("Default bind engine not found.")

        print("Deleting 'orders' bind tables...")
        
        tables_to_drop = [
            OrderProcessing.__table__, 
            Order.__table__,
        ]
        
        db.Model.metadata.drop_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        db.Model.metadata.create_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        
        flash("✅ '주문(Orders)' 테이블이 성공적으로 초기화되었습니다.", "success")

    except Exception as e:
        db.session.rollback()
        print(f"Orders DB Reset Error: {e}")
        traceback.print_exc()
        flash(f"🚨 주문 DB 초기화 중 오류 발생: {e}", "error")
    
    return redirect(url_for('ui.setting_page'))

@api_bp.route('/api/reset-announcements-db', methods=['POST'])
@admin_required
def reset_announcements_db():
    try:
        engine = db.get_engine(bind=None)
        if engine is None:
            raise Exception("Default bind engine not found.")

        print("Deleting 'announcements' bind table...")
        
        tables_to_drop = [Announcement.__table__]
        
        db.Model.metadata.drop_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        db.Model.metadata.create_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        
        flash("✅ '공지사항(Announcements)' 테이블이 성공적으로 초기화되었습니다.", "success")

    except Exception as e:
        db.session.rollback()
        print(f"Announcements DB Reset Error: {e}")
        traceback.print_exc()
        flash(f"🚨 공지사항 DB 초기화 중 오류 발생: {e}", "error")
    
    return redirect(url_for('ui.setting_page'))

@api_bp.route('/api/reset-store-db', methods=['POST'])
@admin_required
def reset_store_db():
    try:
        engine = db.get_engine(bind=None)
        if engine is None:
            raise Exception("Default bind engine not found.")

        print("Deleting 'store_info' bind table...")
        
        # (수정) Staff, ScheduleEvent 테이블도 삭제 리스트에 추가
        tables_to_drop = [
            ScheduleEvent.__table__, 
            Staff.__table__,
            Setting.__table__, 
            User.__table__, 
            Store.__table__, 
            Brand.__table__
        ]
        
        db.Model.metadata.drop_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        db.Model.metadata.create_all(bind=engine, tables=tables_to_drop, checkfirst=True)
        
        flash("✅ '계정/매장/설정/직원/일정' 테이블이 성공적으로 초기화되었습니다. (모든 계정 삭제됨)", "success")

    except Exception as e:
        db.session.rollback()
        print(f"Store Info DB Reset Error: {e}")
        traceback.print_exc()
        flash(f"🚨 계정/매장 DB 초기화 중 오류 발생: {e}", "error")
    
    return redirect(url_for('ui.setting_page'))

# --- (신규) 매장 일정(Schedule) API ---

@api_bp.route('/api/schedule/events', methods=['GET'])
@login_required
def get_schedule_events():
    """ FullCalendar가 요청하는 일정 데이터를 반환합니다. """
    try:
        # FullCalendar가 보내는 start, end 파라미터 (ISO 8601 형식)
        start_str = request.args.get('start')
        end_str = request.args.get('end')

        start_date = _parse_iso_date_string(start_str)
        end_date = _parse_iso_date_string(end_str)

        if not start_date or not end_date:
            return jsonify({'status': 'error', 'message': '날짜 범위가 잘못되었습니다.'}), 400
        
        # (수정) Staff 정보도 함께 로드 (N+1 방지)
        events_query = ScheduleEvent.query.options(
            joinedload(ScheduleEvent.staff) 
        ).filter(
            ScheduleEvent.store_id == current_user.store_id,
            ScheduleEvent.start_time >= start_date,
            ScheduleEvent.start_time < end_date
        )
        
        events = events_query.all()
        
        calendar_events = []
        for event in events:
            # (수정) staff_id가 0이거나 None이면 '매장', 아니면 직원 이름
            staff_name = event.staff.name if event.staff else '매장'
            
            # (수정) FullCalendar 형식에 맞게 데이터 가공
            calendar_events.append({
                'id': event.id,
                'title': f"[{staff_name}] {event.title}",
                'start': event.start_time.isoformat(),
                'end': event.end_time.isoformat() if event.end_time else None,
                'allDay': event.all_day,
                'color': event.color,
                'extendedProps': {
                    'staff_id': event.staff_id or 0,
                    'event_type': event.event_type,
                    'raw_title': event.title
                },
                'classNames': [f'event-type-{event.event_type}'] # (신규) CSS 클래스용
            })
            
        return jsonify(calendar_events)

    except Exception as e:
        print(f"Error fetching schedule events: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/schedule/events', methods=['POST'])
@admin_required
def add_schedule_event():
    """ 새 일정을 등록합니다. """
    data = request.json
    
    try:
        staff_id = int(data.get('staff_id', 0))
        start_date = _parse_iso_date_string(data.get('start_time'))
        end_date = _parse_iso_date_string(data.get('end_time'))
        all_day = bool(data.get('all_day', True))
        title = data.get('title', '').strip()
        event_type = data.get('event_type', '일정').strip()
        color = data.get('color', '#0d6efd')

        if not all([start_date, title, event_type]):
             return jsonify({'status': 'error', 'message': '필수 항목(시작일, 제목, 종류)이 누락되었습니다.'}), 400
        
        # (수정) staff_id가 0이면(매장 전체) null로, 아니면 해당 id로
        final_staff_id = staff_id if staff_id > 0 else None
        
        # (수정) all_day가 True일 때, end_date가 있으면 +1일 처리 (FullCalendar 규칙)
        final_end_time = None
        if not all_day and end_date:
            final_end_time = end_date
        elif all_day and end_date and end_date > start_date:
            # FullCalendar는 allDay 이벤트의 end 날짜를 +1일 해서 줘야 
            # 캘린더 상에 해당 날짜까지 포함된 것으로 표시함
            final_end_time = end_date 

        new_event = ScheduleEvent(
            store_id=current_user.store_id,
            staff_id=final_staff_id,
            title=title,
            event_type=event_type,
            start_time=start_date,
            end_time=final_end_time,
            all_day=all_day,
            color=color
        )
        db.session.add(new_event)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': '일정이 등록되었습니다.', 'event_id': new_event.id}), 201
    
    except Exception as e:
        db.session.rollback()
        print(f"Error adding schedule event: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/schedule/events/<int:event_id>', methods=['POST'])
@admin_required
def update_schedule_event(event_id):
    """ 기존 일정을 수정합니다. (POST로 업데이트 처리) """
    event = ScheduleEvent.query.filter_by(
        id=event_id, 
        store_id=current_user.store_id
    ).first()
    
    if not event:
        return jsonify({'status': 'error', 'message': '수정할 일정을 찾을 수 없습니다.'}), 404
        
    data = request.json
    
    try:
        staff_id = int(data.get('staff_id', 0))
        start_date = _parse_iso_date_string(data.get('start_time'))
        end_date = _parse_iso_date_string(data.get('end_time'))
        all_day = bool(data.get('all_day', True))
        title = data.get('title', '').strip()
        event_type = data.get('event_type', '일정').strip()
        color = data.get('color', '#0d6efd')

        if not all([start_date, title, event_type]):
             return jsonify({'status': 'error', 'message': '필수 항목(시작일, 제목, 종류)이 누락되었습니다.'}), 400
        
        event.staff_id = staff_id if staff_id > 0 else None
        event.title = title
        event.event_type = event_type
        event.start_time = start_date
        event.all_day = all_day
        event.color = color

        final_end_time = None
        if not all_day and end_date:
            final_end_time = end_date
        elif all_day and end_date and end_date > start_date:
            final_end_time = end_date
        event.end_time = final_end_time
        
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': '일정이 수정되었습니다.'})
    
    except Exception as e:
        db.session.rollback()
        print(f"Error updating schedule event: {e}")
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500

@api_bp.route('/api/schedule/events/<int:event_id>', methods=['DELETE'])
@admin_required
def delete_schedule_event(event_id):
    """ 기존 일정을 삭제합니다. """
    try:
        event = ScheduleEvent.query.filter_by(
            id=event_id, 
            store_id=current_user.store_id
        ).first()
        
        if not event:
            return jsonify({'status': 'error', 'message': '삭제할 일정을 찾을 수 없습니다.'}), 404
            
        db.session.delete(event)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': '일정이 삭제되었습니다.'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting schedule event: {e}")
        return jsonify({'status': 'error', 'message': f'서버 오류: {e}'}), 500